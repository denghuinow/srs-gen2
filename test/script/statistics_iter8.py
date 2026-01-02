#!/usr/bin/env python3
"""
统计每个文档的检查项数量、iter_8生成文档使用的语义单元数量、iter_8生成文档的检查项通过数
生成Excel数据表
"""

import re
import os
from pathlib import Path
from typing import Dict, Optional, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def extract_doc_name_from_path(file_path: Path) -> str:
    """从文件路径中提取文档名称（去掉_evaluation.md等后缀）"""
    name = file_path.stem
    # 去掉 _evaluation 后缀
    if name.endswith("_evaluation"):
        name = name[:-11]
    return name


def get_check_item_count_from_eval_md(md_path: Path) -> Optional[int]:
    """从评估报告markdown文件中提取检查项总数"""
    try:
        with open(md_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 查找"检查项总数"行
        pattern = r'检查项总数\s*\|\s*(\d+)'
        match = re.search(pattern, content)
        if match:
            return int(match.group(1))
        
        # 备用模式：查找表格中的检查项总数
        pattern2 = r'\|\s*检查项总数\s*\|\s*(\d+)\s*\|'
        match2 = re.search(pattern2, content)
        if match2:
            return int(match2.group(1))
        
        return None
    except Exception as e:
        print(f"读取 {md_path} 时出错: {e}")
        return None


def count_semantic_units(txt_path: Path) -> Optional[int]:
    """统计语义单元txt文件中的行数（每行一个语义单元）"""
    try:
        with open(txt_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        # 过滤空行
        non_empty_lines = [line.strip() for line in lines if line.strip()]
        return len(non_empty_lines)
    except Exception as e:
        print(f"读取 {txt_path} 时出错: {e}")
        return None


def count_passed_check_items(md_path: Path) -> Optional[int]:
    """从评估报告markdown文件中统计检查项通过数"""
    try:
        with open(md_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 查找"检查项详细结果"表格，统计"✓ 通过"的数量
        # 匹配模式：| 索引 | ... | 多数投票 |
        # 然后查找包含"✓ 通过"的行
        
        # 先找到表格开始位置
        table_start = content.find("## 检查项详细结果")
        if table_start == -1:
            return None
        
        # 提取表格内容
        table_end = content.find("## 数值说明", table_start)
        if table_end == -1:
            table_end = len(content)
        
        table_content = content[table_start:table_end]
        
        # 统计包含"✓ 通过"的行数
        passed_count = len(re.findall(r'✓\s*通过', table_content))
        
        return passed_count
    except Exception as e:
        print(f"读取 {md_path} 时出错: {e}")
        return None


def extract_pool_size_from_log(log_path: Path) -> Optional[int]:
    """从日志文件中提取iter_8的最终需求池大小"""
    try:
        with open(log_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 方法1：查找"pool_iter_8.txt"之前的"需求池更新完成，当前大小"
        pool_iter8_pattern = r'pool_iter_8\.txt'
        pool_iter8_match = re.search(pool_iter8_pattern, content)
        
        if pool_iter8_match:
            # 在pool_iter_8.txt之前查找最近的"需求池更新完成，当前大小"
            before_pool = content[:pool_iter8_match.start()]
            # 从后往前查找最后一个匹配
            pool_size_pattern = r'需求池更新完成，当前大小:\s*(\d+)'
            matches = list(re.finditer(pool_size_pattern, before_pool))
            if matches:
                return int(matches[-1].group(1))
        
        # 方法2：查找"[外循环] 第 8 轮迭代完成"之前的最后一个"需求池更新完成，当前大小"
        iter8_complete_pattern = r'\[外循环\]\s*第\s*8\s*轮迭代完成'
        iter8_complete_match = re.search(iter8_complete_pattern, content)
        
        if iter8_complete_match:
            before_complete = content[:iter8_complete_match.start()]
            pool_size_pattern = r'需求池更新完成，当前大小:\s*(\d+)'
            matches = list(re.finditer(pool_size_pattern, before_complete))
            if matches:
                return int(matches[-1].group(1))
        
        return None
    except Exception as e:
        print(f"读取 {log_path} 时出错: {e}")
        return None


def collect_statistics(
    eval_reports_dir: Path,
    output_dir: Path
) -> List[Dict]:
    """
    收集统计数据
    
    Args:
        eval_reports_dir: 评估报告根目录
        output_dir: 输出目录（用于查找语义单元文件和日志文件）
    
    Returns:
        数据列表，每个元素包含：文档名称、检查项数量、语义单元数量、检查项通过数、需求池数量、语义单元检查项通过数
    """
    d_base_dir = eval_reports_dir / "documents" / "d_base"
    srs_iter8_dir = eval_reports_dir / "documents" / "srs_iter_8"
    units_dir = output_dir / "units_collection" / "pool_iter_8"
    units_eval_dir = eval_reports_dir / "units" / "pool_iter_8"
    
    # 获取所有文档名称（从d_base目录）
    doc_names = set()
    for md_file in d_base_dir.glob("*_evaluation.md"):
        doc_name = extract_doc_name_from_path(md_file)
        doc_names.add(doc_name)
    
    # 收集数据
    data = []
    for doc_name in sorted(doc_names):
        # 1. 检查项数量（从d_base评估报告）
        d_base_md = d_base_dir / f"{doc_name}_evaluation.md"
        check_item_count = get_check_item_count_from_eval_md(d_base_md) if d_base_md.exists() else None
        
        # 2. 语义单元数量（从pool_iter_8的txt文件）
        units_txt = units_dir / f"{doc_name}.txt"
        semantic_units_count = count_semantic_units(units_txt) if units_txt.exists() else None
        
        # 3. 检查项通过数（从srs_iter_8评估报告）
        srs_iter8_md = srs_iter8_dir / f"{doc_name}_evaluation.md"
        passed_count = count_passed_check_items(srs_iter8_md) if srs_iter8_md.exists() else None
        
        # 4. 需求池数量（从日志文件中提取iter_8的需求池大小）
        # 查找日志文件（可能在多个位置）
        log_paths = [
            output_dir / doc_name / "run_attempt_1.log",
            output_dir / doc_name / "run_latest.log",
        ]
        pool_size = None
        for log_path in log_paths:
            if log_path.exists():
                pool_size = extract_pool_size_from_log(log_path)
                if pool_size is not None:
                    break
        
        # 5. 语义单元检查项通过数（从pool_iter_8评估报告）
        units_eval_md = units_eval_dir / f"{doc_name}_evaluation.md"
        units_passed_count = count_passed_check_items(units_eval_md) if units_eval_md.exists() else None
        
        data.append({
            "文档名称": doc_name,
            "检查项数量": check_item_count,
            "语义单元数量": semantic_units_count,
            "检查项通过数": passed_count,
            "需求池数量": pool_size,
            "语义单元检查项通过数": units_passed_count,
        })
    
    return data


def save_to_excel(data: List[Dict], output_path: Path):
    """保存数据到Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "统计数据"
    
    # 设置表头
    headers = ["文档名称", "检查项数量", "语义单元数量", "检查项通过数", "需求池数量", "语义单元检查项通过数"]
    ws.append(headers)
    
    # 设置表头样式
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 写入数据
    for row_data in data:
        ws.append([
            row_data["文档名称"],
            row_data["检查项数量"],
            row_data["语义单元数量"],
            row_data["检查项通过数"],
            row_data["需求池数量"],
            row_data["语义单元检查项通过数"],
        ])
    
    # 设置列宽
    ws.column_dimensions['A'].width = 50  # 文档名称
    ws.column_dimensions['B'].width = 15  # 检查项数量
    ws.column_dimensions['C'].width = 15  # 语义单元数量
    ws.column_dimensions['D'].width = 15  # 检查项通过数
    ws.column_dimensions['E'].width = 15  # 需求池数量
    ws.column_dimensions['F'].width = 20  # 语义单元检查项通过数
    
    # 设置数据对齐
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in [2, 3, 4, 5, 6]:  # B, C, D, E, F列（数字列）
            cell = row[col_idx - 1]
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    wb.save(output_path)
    print(f"Excel文件已保存到: {output_path}")


def main():
    """主函数"""
    # 设置路径
    eval_reports_dir = Path("/root/project/srs/srs-gen2/eval_reports/minimal_en_iter10")
    output_dir = Path("/root/project/srs/srs-gen2/output/minimal_en_iter10")
    output_excel = eval_reports_dir / "iter8统计表.xlsx"
    
    print("开始收集统计数据...")
    data = collect_statistics(eval_reports_dir, output_dir)
    
    print(f"\n共收集到 {len(data)} 个文档的统计数据")
    print("\n前5条数据预览:")
    for i, row in enumerate(data[:5], 1):
        print(f"{i}. {row}")
    
    # 保存到Excel
    print("\n正在保存到Excel...")
    save_to_excel(data, output_excel)
    
    # 显示统计摘要
    check_counts = [d["检查项数量"] for d in data if d["检查项数量"] is not None]
    unit_counts = [d["语义单元数量"] for d in data if d["语义单元数量"] is not None]
    passed_counts = [d["检查项通过数"] for d in data if d["检查项通过数"] is not None]
    pool_sizes = [d["需求池数量"] for d in data if d["需求池数量"] is not None]
    units_passed_counts = [d["语义单元检查项通过数"] for d in data if d["语义单元检查项通过数"] is not None]
    
    print("\n统计摘要:")
    if check_counts:
        print(f"检查项数量 - 平均值: {sum(check_counts)/len(check_counts):.2f}, 范围: {min(check_counts)}-{max(check_counts)}")
    if unit_counts:
        print(f"语义单元数量 - 平均值: {sum(unit_counts)/len(unit_counts):.2f}, 范围: {min(unit_counts)}-{max(unit_counts)}")
    if passed_counts:
        print(f"检查项通过数 - 平均值: {sum(passed_counts)/len(passed_counts):.2f}, 范围: {min(passed_counts)}-{max(passed_counts)}")
    if pool_sizes:
        print(f"需求池数量 - 平均值: {sum(pool_sizes)/len(pool_sizes):.2f}, 范围: {min(pool_sizes)}-{max(pool_sizes)}")
    if units_passed_counts:
        print(f"语义单元检查项通过数 - 平均值: {sum(units_passed_counts)/len(units_passed_counts):.2f}, 范围: {min(units_passed_counts)}-{max(units_passed_counts)}")


if __name__ == "__main__":
    main()

