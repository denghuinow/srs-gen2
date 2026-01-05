#!/usr/bin/env python3
"""
统计每个文档的检查项数量、每个迭代生成文档使用的语义单元数量、检查项通过数
生成Excel数据表（每个迭代一个sheet）
"""

import json
import re
import os
from pathlib import Path
from typing import Dict, Optional, List, Set
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def extract_doc_name_from_path(file_path: Path) -> str:
    """从文件路径中提取文档名称（去掉_evaluation.md等后缀）"""
    name = file_path.stem
    # 去掉 _evaluation 后缀
    if name.endswith("_evaluation"):
        name = name[:-11]
    return name


def get_check_item_count_from_eval_md(md_path: Path) -> Optional[int]:
    """从评估报告markdown文件中提取检查项总数"""
    try:
        with open(md_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 查找"检查项总数"行
        pattern = r'检查项总数\s*\|\s*(\d+)'
        match = re.search(pattern, content)
        if match:
            return int(match.group(1))
        
        # 备用模式：查找表格中的检查项总数
        pattern2 = r'\|\s*检查项总数\s*\|\s*(\d+)\s*\|'
        match2 = re.search(pattern2, content)
        if match2:
            return int(match2.group(1))
        
        return None
    except Exception as e:
        print(f"读取 {md_path} 时出错: {e}")
        return None


def count_semantic_units(txt_path: Path) -> Optional[int]:
    """统计语义单元txt文件中的行数（每行一个语义单元）"""
    try:
        with open(txt_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        # 过滤空行
        non_empty_lines = [line.strip() for line in lines if line.strip()]
        return len(non_empty_lines)
    except Exception as e:
        print(f"读取 {txt_path} 时出错: {e}")
        return None


def count_passed_check_items(md_path: Path) -> Optional[int]:
    """从评估报告markdown文件中统计检查项通过数"""
    try:
        with open(md_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 查找"检查项详细结果"表格，统计"✓ 通过"的数量
        # 匹配模式：| 索引 | ... | 多数投票 |
        # 然后查找包含"✓ 通过"的行
        
        # 先找到表格开始位置
        table_start = content.find("## 检查项详细结果")
        if table_start == -1:
            return None
        
        # 提取表格内容
        table_end = content.find("## 数值说明", table_start)
        if table_end == -1:
            table_end = len(content)
        
        table_content = content[table_start:table_end]
        
        # 统计包含"✓ 通过"的行数
        passed_count = len(re.findall(r'✓\s*通过', table_content))
        
        return passed_count
    except Exception as e:
        print(f"读取 {md_path} 时出错: {e}")
        return None


def load_iter_stats(stats_file: Path) -> Optional[Dict]:
    """从all_iter_stats.json文件加载统计数据"""
    try:
        if not stats_file.exists():
            return None
        with open(stats_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError) as e:
        print(f"读取统计数据文件 {stats_file} 时出错: {e}")
        return None


def get_available_iterations(
    eval_reports_dir: Path,
    output_dir: Path,
    stats_file: Optional[Path] = None
) -> Set[int]:
    """
    获取所有可用的迭代编号
    
    Args:
        eval_reports_dir: 评估报告根目录
        output_dir: 输出目录
        stats_file: 统计数据文件路径（可选）
    
    Returns:
        迭代编号集合
    """
    iterations = set()
    
    # 方法1：从all_iter_stats.json读取
    if stats_file and stats_file.exists():
        stats = load_iter_stats(stats_file)
        if stats and "iterations" in stats:
            for iter_str in stats["iterations"].keys():
                try:
                    iterations.add(int(iter_str))
                except ValueError:
                    pass
    
    # 方法2：从评估报告目录检测
    srs_docs_dir = eval_reports_dir / "documents"
    if srs_docs_dir.exists():
        for subdir in srs_docs_dir.iterdir():
            if subdir.is_dir() and subdir.name.startswith("srs_iter_"):
                try:
                    iter_num = int(subdir.name.replace("srs_iter_", ""))
                    iterations.add(iter_num)
                except ValueError:
                    pass
    
    # 方法3：从units_collection目录检测
    units_collection_dir = output_dir / "units_collection"
    if units_collection_dir.exists():
        for subdir in units_collection_dir.iterdir():
            if subdir.is_dir() and subdir.name.startswith("pool_iter_"):
                try:
                    iter_num = int(subdir.name.replace("pool_iter_", ""))
                    iterations.add(iter_num)
                except ValueError:
                    pass
    
    return sorted(iterations)


def collect_statistics_for_iter(
    eval_reports_dir: Path,
    output_dir: Path,
    iter_num: int,
    all_stats: Optional[Dict] = None  # 保留参数以保持兼容性，但不再使用
) -> List[Dict]:
    """
    为指定迭代收集统计数据
    
    Args:
        eval_reports_dir: 评估报告根目录
        output_dir: 输出目录（用于查找语义单元文件和日志文件）
        iter_num: 迭代编号
        all_stats: 从all_iter_stats.json加载的统计数据（可选）
    
    Returns:
        数据列表，每个元素包含：文档名称、检查项数量、语义单元数量、检查项通过数、需求池数量、语义单元检查项通过数
    """
    d_base_dir = eval_reports_dir / "documents" / "d_base"
    srs_iter_dir = eval_reports_dir / "documents" / f"srs_iter_{iter_num}"
    units_dir = output_dir / "units_collection" / f"pool_iter_{iter_num}"
    units_eval_dir = eval_reports_dir / "units" / f"pool_iter_{iter_num}"
    
    # 获取所有文档名称（从d_base目录）
    doc_names = set()
    for md_file in d_base_dir.glob("*_evaluation.md"):
        doc_name = extract_doc_name_from_path(md_file)
        doc_names.add(doc_name)
    
    # 收集数据
    data = []
    for doc_name in sorted(doc_names):
        # 1. 检查项数量（从d_base评估报告）
        d_base_md = d_base_dir / f"{doc_name}_evaluation.md"
        check_item_count = get_check_item_count_from_eval_md(d_base_md) if d_base_md.exists() else None
        
        # 2. 语义单元数量（从pool_iter_{iter_num}的txt文件读取）
        units_txt = units_dir / f"{doc_name}.txt"
        semantic_units_count = count_semantic_units(units_txt) if units_txt.exists() else None
        
        # 3. 检查项通过数（从srs_iter_{iter_num}评估报告）
        srs_iter_md = srs_iter_dir / f"{doc_name}_evaluation.md"
        passed_count = count_passed_check_items(srs_iter_md) if srs_iter_md.exists() else None
        
        # 4. 需求池数量（从每个文档的all_iter_stats.json读取）
        pool_size = None
        # 尝试从该文档的输出目录读取统计文件
        doc_stats_file = output_dir / doc_name / "all_iter_stats.json"
        doc_stats = load_iter_stats(doc_stats_file)
        if doc_stats and "iterations" in doc_stats:
            iter_stats = doc_stats["iterations"].get(str(iter_num))
            if iter_stats and "pool_size" in iter_stats:
                pool_size = iter_stats["pool_size"]
        
        # 5. 语义单元检查项通过数（从pool_iter_{iter_num}评估报告）
        units_eval_md = units_eval_dir / f"{doc_name}_evaluation.md"
        units_passed_count = count_passed_check_items(units_eval_md) if units_eval_md.exists() else None
        
        data.append({
            "文档名称": doc_name,
            "检查项数量": check_item_count,
            "语义单元数量": semantic_units_count,
            "检查项通过数": passed_count,
            "需求池数量": pool_size,
            "语义单元检查项通过数": units_passed_count,
        })
    
    return data


def collect_statistics(
    eval_reports_dir: Path,
    output_dir: Path,
    iter_nums: Optional[List[int]] = None
) -> Dict[int, List[Dict]]:
    """
    收集多个迭代的统计数据
    
    Args:
        eval_reports_dir: 评估报告根目录
        output_dir: 输出目录（用于查找语义单元文件和日志文件）
        iter_nums: 迭代编号列表，如果为None则自动检测所有迭代
    
    Returns:
        字典，key为迭代编号，value为该迭代的数据列表
    """
    # 加载统计数据文件
    stats_file = output_dir / "all_iter_stats.json"
    all_stats = load_iter_stats(stats_file)
    
    # 如果没有指定迭代编号，自动检测
    if iter_nums is None:
        iter_nums = list(get_available_iterations(eval_reports_dir, output_dir, stats_file))
    
    if not iter_nums:
        print("警告：未找到任何迭代数据")
        return {}
    
    # 为每个迭代收集数据
    result = {}
    for iter_num in sorted(iter_nums):
        print(f"正在收集迭代 {iter_num} 的统计数据...")
        data = collect_statistics_for_iter(eval_reports_dir, output_dir, iter_num, None)
        result[iter_num] = data
    
    return result


def save_to_excel(all_data: Dict[int, List[Dict]], output_path: Path):
    """
    保存数据到Excel文件（每个迭代一个sheet）
    
    Args:
        all_data: 字典，key为迭代编号，value为该迭代的数据列表
        output_path: 输出Excel文件路径
    """
    wb = Workbook()
    # 删除默认的sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # 表头
    headers = ["文档名称", "检查项数量", "语义单元数量", "检查项通过数", "需求池数量", "语义单元检查项通过数"]
    
    # 为每个迭代创建sheet
    for iter_num in sorted(all_data.keys()):
        data = all_data[iter_num]
        
        # 创建sheet（sheet名称不能超过31个字符）
        sheet_name = f"iter_{iter_num}"
        ws = wb.create_sheet(title=sheet_name)
        
        # 设置表头
        ws.append(headers)
        
        # 设置表头样式
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 写入数据
        for row_data in data:
            ws.append([
                row_data["文档名称"],
                row_data["检查项数量"],
                row_data["语义单元数量"],
                row_data["检查项通过数"],
                row_data["需求池数量"],
                row_data["语义单元检查项通过数"],
            ])
        
        # 设置列宽
        ws.column_dimensions['A'].width = 50  # 文档名称
        ws.column_dimensions['B'].width = 15  # 检查项数量
        ws.column_dimensions['C'].width = 15  # 语义单元数量
        ws.column_dimensions['D'].width = 15  # 检查项通过数
        ws.column_dimensions['E'].width = 15  # 需求池数量
        ws.column_dimensions['F'].width = 20  # 语义单元检查项通过数
        
        # 设置数据对齐
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for col_idx in [2, 3, 4, 5, 6]:  # B, C, D, E, F列（数字列）
                cell = row[col_idx - 1]
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    wb.save(output_path)
    print(f"Excel文件已保存到: {output_path}，共 {len(all_data)} 个迭代的统计数据")


def main():
    """主函数"""
    # 设置路径
    eval_reports_dir = Path("/root/project/srs/srs-gen2/eval_reports/minimal_en_iter8")
    output_dir = Path("/root/project/srs/srs-gen2/output/minimal_en_iter8")
    output_excel = eval_reports_dir / "所有迭代统计表.xlsx"
    
    print("开始收集统计数据...")
    
    # 自动检测所有迭代
    stats_file = output_dir / "all_iter_stats.json"
    available_iters = get_available_iterations(eval_reports_dir, output_dir, stats_file)
    
    if not available_iters:
        print("警告：未找到任何迭代数据")
        return
    
    print(f"检测到以下迭代: {sorted(available_iters)}")
    
    # 收集所有迭代的数据
    all_data = collect_statistics(eval_reports_dir, output_dir, list(available_iters))
    
    # 显示每个迭代的统计摘要
    total_docs = 0
    for iter_num in sorted(all_data.keys()):
        data = all_data[iter_num]
        total_docs += len(data)
        print(f"\n迭代 {iter_num} - 共收集到 {len(data)} 个文档的统计数据")
        if data:
            print(f"  前3条数据预览:")
            for i, row in enumerate(data[:3], 1):
                print(f"    {i}. {row['文档名称']}: 检查项={row['检查项数量']}, 语义单元={row['语义单元数量']}, 通过数={row['检查项通过数']}")
    
    # 保存到Excel
    print(f"\n正在保存到Excel（共 {len(all_data)} 个迭代）...")
    save_to_excel(all_data, output_excel)
    
    # 显示所有迭代的统计摘要
    print("\n" + "=" * 60)
    print("所有迭代统计摘要:")
    print("=" * 60)
    
    for iter_num in sorted(all_data.keys()):
        data = all_data[iter_num]
        check_counts = [d["检查项数量"] for d in data if d["检查项数量"] is not None]
        unit_counts = [d["语义单元数量"] for d in data if d["语义单元数量"] is not None]
        passed_counts = [d["检查项通过数"] for d in data if d["检查项通过数"] is not None]
        pool_sizes = [d["需求池数量"] for d in data if d["需求池数量"] is not None]
        units_passed_counts = [d["语义单元检查项通过数"] for d in data if d["语义单元检查项通过数"] is not None]
        
        print(f"\n迭代 {iter_num}:")
        if check_counts:
            print(f"  检查项数量 - 平均值: {sum(check_counts)/len(check_counts):.2f}, 范围: {min(check_counts)}-{max(check_counts)}")
        if unit_counts:
            print(f"  语义单元数量 - 平均值: {sum(unit_counts)/len(unit_counts):.2f}, 范围: {min(unit_counts)}-{max(unit_counts)}")
        if passed_counts:
            print(f"  检查项通过数 - 平均值: {sum(passed_counts)/len(passed_counts):.2f}, 范围: {min(passed_counts)}-{max(passed_counts)}")
        if pool_sizes:
            print(f"  需求池数量 - 平均值: {sum(pool_sizes)/len(pool_sizes):.2f}, 范围: {min(pool_sizes)}-{max(pool_sizes)}")
        if units_passed_counts:
            print(f"  语义单元检查项通过数 - 平均值: {sum(units_passed_counts)/len(units_passed_counts):.2f}, 范围: {min(units_passed_counts)}-{max(units_passed_counts)}")


if __name__ == "__main__":
    main()

