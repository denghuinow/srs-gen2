#!/usr/bin/env python3
"""
统计每个文档的检查项数量、每个迭代生成文档使用的有效需求单元数量、检查项通过数
生成Excel数据表（每个迭代一个sheet）
"""

import json
import re
import os
import ast
from pathlib import Path
from typing import Dict, Optional, List, Set
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def extract_doc_name_from_path(file_path: Path) -> str:
    """从文件路径中提取文档名称（去掉_evaluation.md等后缀）"""
    name = file_path.stem
    # 去掉 _evaluation 后缀
    if name.endswith("_evaluation"):
        name = name[:-11]
    return name


def get_check_item_count_from_eval_md(md_path: Path) -> Optional[int]:
    """从评估报告markdown文件中提取检查项总数"""
    try:
        with open(md_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 查找"检查项总数"行
        pattern = r'检查项总数\s*\|\s*(\d+)'
        match = re.search(pattern, content)
        if match:
            return int(match.group(1))
        
        # 备用模式：查找表格中的检查项总数
        pattern2 = r'\|\s*检查项总数\s*\|\s*(\d+)\s*\|'
        match2 = re.search(pattern2, content)
        if match2:
            return int(match2.group(1))
        
        return None
    except Exception as e:
        print(f"读取 {md_path} 时出错: {e}")
        return None


def count_semantic_units(txt_path: Path) -> Optional[int]:
    """统计语义单元txt文件中的行数（每行一个语义单元）"""
    try:
        with open(txt_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        # 过滤空行
        non_empty_lines = [line.strip() for line in lines if line.strip()]
        return len(non_empty_lines)
    except Exception as e:
        print(f"读取 {txt_path} 时出错: {e}")
        return None


def count_passed_check_items(md_path: Path) -> Optional[int]:
    """从评估报告markdown文件中统计检查项通过数"""
    try:
        with open(md_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 查找"检查项详细结果"表格，统计"✓ 通过"的数量
        # 匹配模式：| 索引 | ... | 多数投票 |
        # 然后查找包含"✓ 通过"的行
        
        # 先找到表格开始位置
        table_start = content.find("## 检查项详细结果")
        if table_start == -1:
            return None
        
        # 提取表格内容
        table_end = content.find("## 数值说明", table_start)
        if table_end == -1:
            table_end = len(content)
        
        table_content = content[table_start:table_end]
        
        # 统计包含"✓ 通过"的行数
        passed_count = len(re.findall(r'✓\s*通过', table_content))
        
        return passed_count
    except Exception as e:
        print(f"读取 {md_path} 时出错: {e}")
        return None


def extract_scores_from_eval_json(json_path: Path) -> Dict[str, Optional[float]]:
    """从评估报告JSON文件中提取加权分、投票通过率、平均通过率"""
    try:
        if not json_path.exists():
            return {"weighted_score": None, "voting_score": None, "average_score": None}
        
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        scores = data.get("scores", {})
        return {
            "weighted_score": scores.get("weighted_score"),
            "voting_score": scores.get("voting_score"),
            "average_score": scores.get("average_score"),
        }
    except Exception as e:
        print(f"读取 {json_path} 时出错: {e}")
        return {"weighted_score": None, "voting_score": None, "average_score": None}


def extract_title_from_titles_json(titles_json_path: Path, doc_name: str) -> Optional[str]:
    """从titles.json文件中读取文档标题"""
    try:
        if not titles_json_path.exists():
            return None
        
        with open(titles_json_path, 'r', encoding='utf-8') as f:
            titles_data = json.load(f)
        
        # 文档名称格式可能是 "0000 - cctns.pdf" 或 "0000 - cctns.pdf.md"
        # titles.json中的key格式是 "0000 - cctns.pdf.md"
        key = doc_name
        if not key.endswith('.md'):
            key = f"{doc_name}.md"
        
        return titles_data.get(key)
    except Exception as e:
        print(f"从titles.json读取标题时出错: {e}")
        return None


def extract_summary_from_r_base(r_base_path: Path) -> Optional[str]:
    """从r_base文本文件中提取完整内容"""
    try:
        if not r_base_path.exists():
            return None
        
        with open(r_base_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 返回完整内容，去除首尾空白
        summary = content.strip()
        return summary if summary else None
    except Exception as e:
        print(f"从r_base读取摘要时出错: {e}")
        return None


def load_iter_stats(stats_file: Path) -> Optional[Dict]:
    """从all_iter_stats.json文件加载统计数据"""
    try:
        if not stats_file.exists():
            return None
        with open(stats_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError) as e:
        print(f"读取统计数据文件 {stats_file} 时出错: {e}")
        return None


def parse_log_for_d_base(log_path: Path) -> Optional[Dict]:
    """
    从日志文件中解析 d_base 的数据
    
    Args:
        log_path: 日志文件路径
    
    Returns:
        包含 pool_size, semantic_units_count, invalid_rate 的字典，如果解析失败返回 None
    """
    try:
        if not log_path.exists():
            return None
        
        with open(log_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 提取需求池大小：匹配 "需求池初始化完成，当前大小: X"
        pool_size_match = re.search(r'需求池初始化完成，当前大小:\s*(\d+)', content)
        if not pool_size_match:
            return None
        pool_size = int(pool_size_match.group(1))
        
        # 提取评分分布：匹配 "评分完成，分布: {...}"
        # 注意：这里可能有多个评分完成，我们需要找到第一个（d_base的）
        # 查找在"需求池初始化完成"之前的评分完成
        pool_init_pos = content.find("需求池初始化完成")
        if pool_init_pos == -1:
            return None
        
        # 在需求池初始化之前查找评分完成
        before_pool_init = content[:pool_init_pos]
        # 匹配评分分布，支持嵌套的字典格式（如 {-2: 186, 1: 1}）
        grade_dist_match = re.search(r'评分完成，分布:\s*(\{[^}]*\})', before_pool_init)
        if not grade_dist_match:
            return None
        
        # 解析评分分布（Python dict格式）
        grade_dist_str = grade_dist_match.group(1)
        try:
            grade_dist = ast.literal_eval(grade_dist_str)
        except (ValueError, SyntaxError):
            return None
        
        # 计算 grade > 0 的单元数
        valid_count = 0
        for grade, count in grade_dist.items():
            if isinstance(grade, (int, float)) and grade > 0:
                valid_count += count
        
        # 计算无效率
        invalid_rate = None
        if pool_size > 0:
            invalid_rate = (pool_size - valid_count) / pool_size * 100
        
        return {
            "pool_size": pool_size,
            "semantic_units_count": valid_count,
            "invalid_rate": invalid_rate,
        }
    except Exception as e:
        print(f"解析日志文件 {log_path} 时出错: {e}")
        return None


def parse_log_for_no_explore_clarify(log_path: Path, d_base_valid_count: int) -> Optional[Dict]:
    """
    从日志文件中解析 no-explore-clarify 的数据
    
    Args:
        log_path: 日志文件路径
        d_base_valid_count: d_base 澄清后 grade > 0 的单元数
    
    Returns:
        包含 semantic_units_count, pool_size, invalid_rate 的字典，如果解析失败返回 None
    """
    try:
        if not log_path.exists():
            return None
        
        with open(log_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 提取基线单元总数：匹配 "基线 SRS 拆分为 X 个语义单元"
        baseline_match = re.search(r'基线 SRS 拆分为\s*(\d+)\s*个语义单元', content)
        if not baseline_match:
            return None
        total_count = int(baseline_match.group(1))
        
        # 使用 d_base 的有效单元数作为有效需求单元数量
        # 计算无效率
        invalid_rate = None
        if total_count > 0:
            invalid_rate = (total_count - d_base_valid_count) / total_count * 100
        
        return {
            "semantic_units_count": d_base_valid_count,
            "pool_size": total_count,
            "invalid_rate": invalid_rate,
        }
    except Exception as e:
        print(f"解析日志文件 {log_path} 时出错: {e}")
        return None


def parse_log_for_no_clarify(log_path: Path, d_base_valid_count: int) -> Optional[Dict]:
    """
    从日志文件中解析 no-clarify 的数据
    
    Args:
        log_path: 日志文件路径
        d_base_valid_count: d_base 澄清后 grade > 0 的单元数
    
    Returns:
        包含 semantic_units_count, pool_size, invalid_rate 的字典，如果解析失败返回 None
    """
    try:
        if not log_path.exists():
            return None
        
        with open(log_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 提取单元总数：匹配 "已保存 X 个语义单元到: .../no-clarify.txt"
        no_clarify_match = re.search(r'已保存\s*(\d+)\s*个语义单元到:.*no-clarify\.txt', content)
        if not no_clarify_match:
            return None
        total_count = int(no_clarify_match.group(1))
        
        # 使用 d_base 的有效单元数作为有效需求单元数量
        # 计算无效率
        invalid_rate = None
        if total_count > 0:
            invalid_rate = (total_count - d_base_valid_count) / total_count * 100
        
        return {
            "semantic_units_count": d_base_valid_count,
            "pool_size": total_count,
            "invalid_rate": invalid_rate,
        }
    except Exception as e:
        print(f"解析日志文件 {log_path} 时出错: {e}")
        return None


def get_available_iterations(
    eval_reports_dir: Path,
    output_dir: Path,
    stats_file: Optional[Path] = None
) -> Set[int]:
    """
    获取所有可用的迭代编号
    
    Args:
        eval_reports_dir: 评估报告根目录
        output_dir: 输出目录
        stats_file: 统计数据文件路径（可选）
    
    Returns:
        迭代编号集合
    """
    iterations = set()
    
    # 方法1：从all_iter_stats.json读取
    if stats_file and stats_file.exists():
        stats = load_iter_stats(stats_file)
        if stats and "iterations" in stats:
            for iter_str in stats["iterations"].keys():
                try:
                    iterations.add(int(iter_str))
                except ValueError:
                    pass
    
    # 方法2：从评估报告目录检测
    srs_docs_dir = eval_reports_dir / "documents"
    if srs_docs_dir.exists():
        for subdir in srs_docs_dir.iterdir():
            if subdir.is_dir() and subdir.name.startswith("srs_iter_"):
                try:
                    iter_num = int(subdir.name.replace("srs_iter_", ""))
                    iterations.add(iter_num)
                except ValueError:
                    pass
    
    # 方法3：从units_collection目录检测
    units_collection_dir = output_dir / "units_collection"
    if units_collection_dir.exists():
        for subdir in units_collection_dir.iterdir():
            if subdir.is_dir() and subdir.name.startswith("pool_iter_"):
                try:
                    iter_num = int(subdir.name.replace("pool_iter_", ""))
                    iterations.add(iter_num)
                except ValueError:
                    pass
    
    return sorted(iterations)


def collect_statistics_for_iter(
    eval_reports_dir: Path,
    output_dir: Path,
    iter_num: int,
    all_stats: Optional[Dict] = None  # 保留参数以保持兼容性，但不再使用
) -> List[Dict]:
    """
    为指定迭代收集统计数据
    
    Args:
        eval_reports_dir: 评估报告根目录
        output_dir: 输出目录（用于查找语义单元文件和日志文件）
        iter_num: 迭代编号
        all_stats: 从all_iter_stats.json加载的统计数据（可选）
    
    Returns:
        数据列表，每个元素包含：文档名称、检查项数量、有效需求单元数量、SRS检查项通过数、需求池数量、语义单元检查项通过数、无效需求率、SRS阶段合并通过数、语义单元阶段合并通过数
    """
    d_base_dir = eval_reports_dir / "documents" / "d_base"
    srs_iter_dir = eval_reports_dir / "documents" / f"srs_iter_{iter_num}"
    srs_merge_dir = eval_reports_dir / "documents" / f"srs_iter_1_to_{iter_num}"
    units_dir = output_dir / "units_collection" / f"pool_iter_{iter_num}"
    units_eval_dir = eval_reports_dir / "units" / f"pool_iter_{iter_num}"
    units_merge_dir = eval_reports_dir / "units" / f"pool_iter_1_to_{iter_num}"
    
    # 获取所有文档名称（从d_base目录）
    doc_names = set()
    for md_file in d_base_dir.glob("*_evaluation.md"):
        doc_name = extract_doc_name_from_path(md_file)
        doc_names.add(doc_name)
    
    # 收集数据
    data = []
    for doc_name in sorted(doc_names):
        # 1. 检查项数量（从d_base评估报告）
        d_base_md = d_base_dir / f"{doc_name}_evaluation.md"
        check_item_count = get_check_item_count_from_eval_md(d_base_md) if d_base_md.exists() else None
        
        # 2. 有效需求单元数量（从pool_iter_{iter_num}的txt文件读取）
        units_txt = units_dir / f"{doc_name}.txt"
        semantic_units_count = count_semantic_units(units_txt) if units_txt.exists() else None
        
        # 3. SRS检查项通过数（从srs_iter_{iter_num}评估报告）
        srs_iter_md = srs_iter_dir / f"{doc_name}_evaluation.md"
        srs_passed_count = count_passed_check_items(srs_iter_md) if srs_iter_md.exists() else None
        
        # 4. 需求池数量（从每个文档的all_iter_stats.json读取）
        pool_size = None
        # 尝试从该文档的输出目录读取统计文件
        doc_stats_file = output_dir / doc_name / "all_iter_stats.json"
        doc_stats = load_iter_stats(doc_stats_file)
        if doc_stats and "iterations" in doc_stats:
            iter_stats = doc_stats["iterations"].get(str(iter_num))
            if iter_stats and "pool_size" in iter_stats:
                pool_size = iter_stats["pool_size"]
        
        # 5. 语义单元检查项通过数（从pool_iter_{iter_num}评估报告）
        units_eval_md = units_eval_dir / f"{doc_name}_evaluation.md"
        units_passed_count = count_passed_check_items(units_eval_md) if units_eval_md.exists() else None
        
        # 6. 无效需求率 = (需求池数量 - 有效需求单元数量) / 需求池数量 * 100
        invalid_rate = None
        if pool_size is not None and pool_size > 0 and semantic_units_count is not None:
            invalid_rate = (pool_size - semantic_units_count) / pool_size * 100
        
        # 7. SRS阶段合并通过数（从srs_iter_1_to_{iter_num}评估报告）
        srs_merge_md = srs_merge_dir / f"{doc_name}_evaluation.md"
        srs_merge_passed_count = count_passed_check_items(srs_merge_md) if srs_merge_md.exists() else None
        
        # 8. 语义单元阶段合并通过数（从pool_iter_1_to_{iter_num}评估报告）
        units_merge_md = units_merge_dir / f"{doc_name}_evaluation.md"
        units_merge_passed_count = count_passed_check_items(units_merge_md) if units_merge_md.exists() else None
        
        data.append({
            "doc_name": doc_name,
            "check_item_count": check_item_count,
            "valid_semantic_units_count": semantic_units_count,
            "pool_size": pool_size,
            "srs_passed_count": srs_passed_count,
            "srs_merge_passed_count": srs_merge_passed_count,
            "units_passed_count": units_passed_count,
            "units_merge_passed_count": units_merge_passed_count,
            "invalid_rate": invalid_rate,
        })
    
    return data


def collect_statistics(
    eval_reports_dir: Path,
    output_dir: Path,
    iter_nums: Optional[List[int]] = None
) -> Dict[int, List[Dict]]:
    """
    收集多个迭代的统计数据
    
    Args:
        eval_reports_dir: 评估报告根目录
        output_dir: 输出目录（用于查找语义单元文件和日志文件）
        iter_nums: 迭代编号列表，如果为None则自动检测所有迭代
    
    Returns:
        字典，key为迭代编号，value为该迭代的数据列表
    """
    # 加载统计数据文件
    stats_file = output_dir / "all_iter_stats.json"
    all_stats = load_iter_stats(stats_file)
    
    # 如果没有指定迭代编号，自动检测
    if iter_nums is None:
        iter_nums = list(get_available_iterations(eval_reports_dir, output_dir, stats_file))
    
    if not iter_nums:
        print("警告：未找到任何迭代数据")
        return {}
    
    # 为每个迭代收集数据
    result = {}
    for iter_num in sorted(iter_nums):
        print(f"正在收集迭代 {iter_num} 的统计数据...")
        data = collect_statistics_for_iter(eval_reports_dir, output_dir, iter_num, None)
        result[iter_num] = data
    
    return result


def collect_statistics_for_special_versions(
    eval_reports_dir: Path,
    output_dir: Path
) -> Dict[str, List[Dict]]:
    """
    收集 d_base、no-explore-clarify、no-clarify 的统计数据
    
    Args:
        eval_reports_dir: 评估报告根目录
        output_dir: 输出目录（用于查找日志文件）
    
    Returns:
        字典，key为版本名称（"d_base", "no-explore-clarify", "no-clarify"），value为该版本的数据列表
    """
    d_base_dir = eval_reports_dir / "documents" / "d_base"
    
    # 获取所有文档名称（从d_base目录）
    doc_names = set()
    for md_file in d_base_dir.glob("*_evaluation.md"):
        doc_name = extract_doc_name_from_path(md_file)
        doc_names.add(doc_name)
    
    result = {
        "d_base": [],
        "no-explore-clarify": [],
        "no-clarify": [],
    }
    
    for doc_name in sorted(doc_names):
        # 查找日志文件
        log_path = output_dir / doc_name / "run_latest.log"
        
        # 先解析 d_base 数据
        d_base_data = parse_log_for_d_base(log_path)
        if not d_base_data:
            # 如果无法解析 d_base，跳过该文档
            continue
        
        d_base_valid_count = d_base_data["semantic_units_count"]
        
        # 解析 no-explore-clarify 数据
        no_explore_clarify_data = parse_log_for_no_explore_clarify(log_path, d_base_valid_count)
        
        # 解析 no-clarify 数据
        no_clarify_data = parse_log_for_no_clarify(log_path, d_base_valid_count)
        
        # 添加到结果中
        result["d_base"].append({
            "doc_name": doc_name,
            "pool_size": d_base_data["pool_size"],
            "valid_semantic_units_count": d_base_data["semantic_units_count"],
            "invalid_rate": d_base_data["invalid_rate"],
        })
        
        if no_explore_clarify_data:
            result["no-explore-clarify"].append({
                "doc_name": doc_name,
                "pool_size": no_explore_clarify_data["pool_size"],
                "valid_semantic_units_count": no_explore_clarify_data["semantic_units_count"],
                "invalid_rate": no_explore_clarify_data["invalid_rate"],
            })
        
        if no_clarify_data:
            result["no-clarify"].append({
                "doc_name": doc_name,
                "pool_size": no_clarify_data["pool_size"],
                "valid_semantic_units_count": no_clarify_data["semantic_units_count"],
                "invalid_rate": no_clarify_data["invalid_rate"],
            })
    
    return result


def create_pass_rate_segmentation_sheet(
    wb: Workbook,
    all_data: Dict[int, List[Dict]],
    eval_reports_dir: Path,
    output_dir: Path,
    max_iter: int,
    sheet_index: Optional[int] = None
):
    """
    创建按通过率分段的汇总sheet
    
    Args:
        wb: Excel工作簿
        all_data: 所有迭代的数据
        eval_reports_dir: 评估报告根目录
        output_dir: 输出目录
        max_iter: 最大迭代编号（使用srs_iter_1_to_{max_iter}的合并阶段数据）
    """
    # 使用最后一个迭代的数据获取文档列表和检查项数量
    if max_iter not in all_data:
        return
    
    data = all_data[max_iter]
    srs_merge_dir = eval_reports_dir / "documents" / f"srs_iter_1_to_{max_iter}"
    r_base_dir = output_dir / "srs_collection" / "r_base"
    titles_json_path = Path("/root/project/srs/srs-docs/resources/titles.json")
    
    # 收集所有文档的详细信息
    doc_details = []
    for row_data in data:
        doc_name = row_data["doc_name"]
        check_item_count = row_data["check_item_count"]
        srs_merge_passed_count = row_data["srs_merge_passed_count"]
        invalid_rate = row_data["invalid_rate"]
        pool_size = row_data["pool_size"]
        
        # 计算通过率
        pass_rate = None
        if check_item_count is not None and check_item_count > 0 and srs_merge_passed_count is not None:
            pass_rate = (srs_merge_passed_count / check_item_count) * 100
        
        # 从合并阶段评估报告JSON中提取加权分、投票通过率
        eval_json = srs_merge_dir / f"{doc_name}_evaluation.json"
        scores = extract_scores_from_eval_json(eval_json)
        weighted_score = scores.get("weighted_score")
        voting_score = scores.get("voting_score")
        
        # 从titles.json读取标题
        title = extract_title_from_titles_json(titles_json_path, doc_name)
        
        # 从r_base文本读取摘要
        r_base_path = r_base_dir / f"{doc_name}.md"
        summary = extract_summary_from_r_base(r_base_path)
        
        doc_details.append({
            "doc_name": doc_name,
            "title": title or "",
            "summary": summary or "",
            "weighted_score": weighted_score,
            "voting_score": voting_score,
            "pass_rate": pass_rate,
            "pool_size": pool_size,
            "invalid_rate": invalid_rate,
        })
    
    # 过滤掉没有通过率的数据
    valid_docs = [d for d in doc_details if d["pass_rate"] is not None]
    if not valid_docs:
        return
    
    # 按通过率排序
    valid_docs.sort(key=lambda x: x["pass_rate"], reverse=True)
    
    # 分成3段
    total_docs = len(valid_docs)
    segment_size = total_docs // 3
    remainder = total_docs % 3
    
    # 分配文档到3段（尽量平均分配）
    segments = []
    start_idx = 0
    for i in range(3):
        # 前remainder个段多分配一个文档
        size = segment_size + (1 if i < remainder else 0)
        end_idx = start_idx + size
        segments.append(valid_docs[start_idx:end_idx])
        start_idx = end_idx
    
    # 创建sheet
    sheet_name = "Pass Rate Segmentation"
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    if sheet_index is not None:
        ws = wb.create_sheet(title=sheet_name, index=sheet_index)
    else:
        ws = wb.create_sheet(title=sheet_name)
    
    # 表头
    headers = ["Doc Name", "SRS Title", "Summary", "Weighted Score", "Voting Pass Rate", "Pass Rate", "Pool Size", "Invalid Rate"]
    ws.append(headers)
    
    # 设置表头样式
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 写入各段数据
    segment_names = ["High Pass Rate", "Medium Pass Rate", "Low Pass Rate"]
    current_row = 2
    
    for seg_idx, (segment_name, segment_docs) in enumerate(zip(segment_names, segments)):
        if not segment_docs:
            continue
        
        # 写入段标题
        ws.cell(row=current_row, column=1, value=f"【{segment_name}】")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
        current_row += 1
        
        # 写入该段的文档数据
        for doc in segment_docs:
            ws.append([
                doc["doc_name"],
                doc["title"],
                doc["summary"],
                round(doc["weighted_score"], 2) if doc["weighted_score"] is not None else None,
                round(doc["voting_score"], 2) if doc["voting_score"] is not None else None,
                round(doc["pass_rate"], 2) if doc["pass_rate"] is not None else None,
                doc["pool_size"],
                round(doc["invalid_rate"], 2) if doc["invalid_rate"] is not None else None,
            ])
            current_row += 1
        
        # 计算并写入该段的平均值
        avg_weighted = sum(d["weighted_score"] for d in segment_docs if d["weighted_score"] is not None)
        avg_voting = sum(d["voting_score"] for d in segment_docs if d["voting_score"] is not None)
        avg_pass_rate = sum(d["pass_rate"] for d in segment_docs if d["pass_rate"] is not None)
        avg_pool_size = sum(d["pool_size"] for d in segment_docs if d["pool_size"] is not None)
        avg_invalid_rate = sum(d["invalid_rate"] for d in segment_docs if d["invalid_rate"] is not None)
        
        count_weighted = sum(1 for d in segment_docs if d["weighted_score"] is not None)
        count_voting = sum(1 for d in segment_docs if d["voting_score"] is not None)
        count_pass_rate = sum(1 for d in segment_docs if d["pass_rate"] is not None)
        count_pool_size = sum(1 for d in segment_docs if d["pool_size"] is not None)
        count_invalid_rate = sum(1 for d in segment_docs if d["invalid_rate"] is not None)
        
        ws.append([
            f"平均值",
            "",
            "",
            round(avg_weighted / count_weighted, 2) if count_weighted > 0 else None,
            round(avg_voting / count_voting, 2) if count_voting > 0 else None,
            round(avg_pass_rate / count_pass_rate, 2) if count_pass_rate > 0 else None,
            round(avg_pool_size / count_pool_size, 2) if count_pool_size > 0 else None,
            round(avg_invalid_rate / count_invalid_rate, 2) if count_invalid_rate > 0 else None,
        ])
        ws.cell(row=current_row, column=1).font = Font(bold=True, italic=True)
        current_row += 1
        
        # 空行分隔
        current_row += 1
    
    # 设置列宽
    ws.column_dimensions['A'].width = 50  # 文档名称
    ws.column_dimensions['B'].width = 40  # SRS标题
    ws.column_dimensions['C'].width = 60  # 摘要内容
    ws.column_dimensions['D'].width = 12  # 加权分
    ws.column_dimensions['E'].width = 15  # 投票通过率
    ws.column_dimensions['F'].width = 12  # 通过率
    ws.column_dimensions['G'].width = 15  # 需求池数量
    ws.column_dimensions['H'].width = 12  # 无效率
    
    # 设置数据对齐
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in [4, 5, 6, 7, 8]:  # D, E, F, G, H列（数字列）
            cell = row[col_idx - 1]
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # A, B, C列左对齐
        for col_idx in [1, 2, 3]:
            cell = row[col_idx - 1]
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)


def create_invalid_rate_segmentation_sheet(
    wb: Workbook,
    all_data: Dict[int, List[Dict]],
    eval_reports_dir: Path,
    output_dir: Path,
    max_iter: int,
    sheet_index: Optional[int] = None
):
    """
    创建按无效率分段的汇总sheet
    
    Args:
        wb: Excel工作簿
        all_data: 所有迭代的数据
        eval_reports_dir: 评估报告根目录
        output_dir: 输出目录
        max_iter: 最大迭代编号（使用srs_iter_1_to_{max_iter}的合并阶段数据）
    """
    # 使用最后一个迭代的数据获取文档列表和无效需求率
    if max_iter not in all_data:
        return
    
    data = all_data[max_iter]
    srs_merge_dir = eval_reports_dir / "documents" / f"srs_iter_1_to_{max_iter}"
    r_base_dir = output_dir / "srs_collection" / "r_base"
    titles_json_path = Path("/root/project/srs/srs-docs/resources/titles.json")
    
    # 收集所有文档的详细信息
    doc_details = []
    for row_data in data:
        doc_name = row_data["doc_name"]
        check_item_count = row_data["check_item_count"]
        srs_merge_passed_count = row_data["srs_merge_passed_count"]
        invalid_rate = row_data["invalid_rate"]
        pool_size = row_data["pool_size"]
        
        # 计算通过率
        pass_rate = None
        if check_item_count is not None and check_item_count > 0 and srs_merge_passed_count is not None:
            pass_rate = (srs_merge_passed_count / check_item_count) * 100
        
        # 从合并阶段评估报告JSON中提取加权分、投票通过率
        eval_json = srs_merge_dir / f"{doc_name}_evaluation.json"
        scores = extract_scores_from_eval_json(eval_json)
        weighted_score = scores.get("weighted_score")
        voting_score = scores.get("voting_score")
        
        # 从titles.json读取标题
        title = extract_title_from_titles_json(titles_json_path, doc_name)
        
        # 从r_base文本读取摘要
        r_base_path = r_base_dir / f"{doc_name}.md"
        summary = extract_summary_from_r_base(r_base_path)
        
        doc_details.append({
            "doc_name": doc_name,
            "title": title or "",
            "summary": summary or "",
            "weighted_score": weighted_score,
            "voting_score": voting_score,
            "pass_rate": pass_rate,
            "pool_size": pool_size,
            "invalid_rate": invalid_rate,
        })
    
    # 过滤掉没有无效率的数据
    valid_docs = [d for d in doc_details if d["invalid_rate"] is not None]
    if not valid_docs:
        return
    
    # 按无效率排序（从高到低）
    valid_docs.sort(key=lambda x: x["invalid_rate"], reverse=True)
    
    # 分成3段
    total_docs = len(valid_docs)
    segment_size = total_docs // 3
    remainder = total_docs % 3
    
    # 分配文档到3段（尽量平均分配）
    segments = []
    start_idx = 0
    for i in range(3):
        # 前remainder个段多分配一个文档
        size = segment_size + (1 if i < remainder else 0)
        end_idx = start_idx + size
        segments.append(valid_docs[start_idx:end_idx])
        start_idx = end_idx
    
    # 创建sheet
    sheet_name = "Invalid Rate Segmentation"
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    if sheet_index is not None:
        ws = wb.create_sheet(title=sheet_name, index=sheet_index)
    else:
        ws = wb.create_sheet(title=sheet_name)
    
    # 表头
    headers = ["Doc Name", "SRS Title", "Summary", "Weighted Score", "Voting Pass Rate", "Pass Rate", "Pool Size", "Invalid Rate"]
    ws.append(headers)
    
    # 设置表头样式
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 写入各段数据
    segment_names = ["High Invalid Rate", "Medium Invalid Rate", "Low Invalid Rate"]
    current_row = 2
    
    for seg_idx, (segment_name, segment_docs) in enumerate(zip(segment_names, segments)):
        if not segment_docs:
            continue
        
        # 写入段标题
        ws.cell(row=current_row, column=1, value=f"【{segment_name}】")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
        current_row += 1
        
        # 写入该段的文档数据
        for doc in segment_docs:
            ws.append([
                doc["doc_name"],
                doc["title"],
                doc["summary"],
                round(doc["weighted_score"], 2) if doc["weighted_score"] is not None else None,
                round(doc["voting_score"], 2) if doc["voting_score"] is not None else None,
                round(doc["pass_rate"], 2) if doc["pass_rate"] is not None else None,
                doc["pool_size"],
                round(doc["invalid_rate"], 2) if doc["invalid_rate"] is not None else None,
            ])
            current_row += 1
        
        # 计算并写入该段的平均值
        avg_weighted = sum(d["weighted_score"] for d in segment_docs if d["weighted_score"] is not None)
        avg_voting = sum(d["voting_score"] for d in segment_docs if d["voting_score"] is not None)
        avg_pass_rate = sum(d["pass_rate"] for d in segment_docs if d["pass_rate"] is not None)
        avg_pool_size = sum(d["pool_size"] for d in segment_docs if d["pool_size"] is not None)
        avg_invalid_rate = sum(d["invalid_rate"] for d in segment_docs if d["invalid_rate"] is not None)
        
        count_weighted = sum(1 for d in segment_docs if d["weighted_score"] is not None)
        count_voting = sum(1 for d in segment_docs if d["voting_score"] is not None)
        count_pass_rate = sum(1 for d in segment_docs if d["pass_rate"] is not None)
        count_pool_size = sum(1 for d in segment_docs if d["pool_size"] is not None)
        count_invalid_rate = sum(1 for d in segment_docs if d["invalid_rate"] is not None)
        
        ws.append([
            f"平均值",
            "",
            "",
            round(avg_weighted / count_weighted, 2) if count_weighted > 0 else None,
            round(avg_voting / count_voting, 2) if count_voting > 0 else None,
            round(avg_pass_rate / count_pass_rate, 2) if count_pass_rate > 0 else None,
            round(avg_pool_size / count_pool_size, 2) if count_pool_size > 0 else None,
            round(avg_invalid_rate / count_invalid_rate, 2) if count_invalid_rate > 0 else None,
        ])
        ws.cell(row=current_row, column=1).font = Font(bold=True, italic=True)
        current_row += 1
        
        # 空行分隔
        current_row += 1
    
    # 设置列宽
    ws.column_dimensions['A'].width = 50  # 文档名称
    ws.column_dimensions['B'].width = 40  # SRS标题
    ws.column_dimensions['C'].width = 60  # 摘要内容
    ws.column_dimensions['D'].width = 12  # 加权分
    ws.column_dimensions['E'].width = 15  # 投票通过率
    ws.column_dimensions['F'].width = 12  # 通过率
    ws.column_dimensions['G'].width = 15  # 需求池数量
    ws.column_dimensions['H'].width = 12  # 无效率
    
    # 设置数据对齐
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in [4, 5, 6, 7, 8]:  # D, E, F, G, H列（数字列）
            cell = row[col_idx - 1]
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # A, B, C列左对齐
        for col_idx in [1, 2, 3]:
            cell = row[col_idx - 1]
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)


def create_weight_config_sheet(
    wb: Workbook,
    sheet_index: Optional[int] = None
):
    """
    创建权重配置表sheet
    
    Args:
        wb: Excel工作簿
        sheet_index: sheet插入位置（可选）
    """
    # 权重配置数据
    weights_config = {
        "FUNCTIONAL": {
            "name_en": "Functional Coverage / Behavioral Rules",
            "weight": 0.25,
        },
        "BUSINESS_FLOW": {
            "name_en": "Business Flow Completeness",
            "weight": 0.15,
        },
        "BOUNDARY": {
            "name_en": "Boundary Condition Completeness",
            "weight": 0.10,
        },
        "EXCEPTION": {
            "name_en": "Exception Handling Coverage",
            "weight": 0.20,
        },
        "DATA_STATE": {
            "name_en": "Data and State Completeness",
            "weight": 0.20,
        },
        "CONSISTENCY_RULE": {
            "name_en": "Consistency / Conflict Detection",
            "weight": 0.10,
        },
    }
    
    # 缩放系数
    scale_factor = 2.0
    
    # 创建sheet
    sheet_name = "Weight Configuration"
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    if sheet_index is not None:
        ws = wb.create_sheet(title=sheet_name, index=sheet_index)
    else:
        ws = wb.create_sheet(title=sheet_name)
    
    # 表头
    headers = ["Dimension Code", "Dimension Name", "Weight Value", "Weight Percentage (%)"]
    ws.append(headers)
    
    # 设置表头样式
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 写入权重数据
    for dim_code, config in weights_config.items():
        weight_value = config["weight"]
        weight_percent = weight_value * 100
        ws.append([
            dim_code,
            config["name_en"],
            weight_value,
            round(weight_percent, 2),
        ])
    
    # 添加汇总行
    ws.append([])  # 空行
    ws.append(["Total Weight", "", sum(c["weight"] for c in weights_config.values()), "100.00"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=3).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=4).font = Font(bold=True)
    
    # 添加计算公式说明
    ws.append([])  # 空行
    ws.append(["Calculation Formula", "", "", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11)
    
    ws.append(["", "Step 1: Calculate Weighted Average", "", ""])
    ws.append(["", "Weighted Average = Σ(Dimension Score × Dimension Weight)", "", ""])
    ws.cell(row=ws.max_row, column=2).font = Font(italic=True)
    
    ws.append([])  # 空行
    ws.append(["", "Step 2: Multiply by Scale Factor", "", ""])
    ws.append(["", f"Weighted Score = Weighted Average × {scale_factor}", "", ""])
    ws.cell(row=ws.max_row, column=2).font = Font(italic=True)
    
    ws.append([])  # 空行
    ws.append(["", "Step 3: Apply Upper Limit", "", ""])
    ws.append(["", "Final Score = min(100.0, Weighted Score)", "", ""])
    ws.cell(row=ws.max_row, column=2).font = Font(italic=True)
    
    ws.append([])  # 空行
    ws.append(["Scale Factor", f"{scale_factor}", "", ""])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 25  # 维度代码
    ws.column_dimensions['B'].width = 30  # 维度名称
    ws.column_dimensions['C'].width = 15  # 权重值
    ws.column_dimensions['D'].width = 18  # 权重百分比
    
    # 设置数据对齐
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in range(1, 5):  # A, B, C, D列
            cell = row[col_idx - 1]
            if col_idx in [1, 2]:  # A, B列（维度代码和名称）左对齐
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:  # C, D列（数字列）居中
                cell.alignment = Alignment(horizontal='center', vertical='center')


def save_to_excel(all_data: Dict[int, List[Dict]], output_path: Path, eval_reports_dir: Path, output_dir: Path, source_excel_path: Path, special_versions_data: Optional[Dict[str, List[Dict]]] = None):
    """
    保存数据到Excel文件（每个迭代一个sheet），基于现有Excel文件追加sheet
    
    Args:
        all_data: 字典，key为迭代编号，value为该迭代的数据列表
        output_path: 输出Excel文件路径
        eval_reports_dir: 评估报告根目录
        output_dir: 输出目录
        source_excel_path: 源Excel文件路径（数据汇总v2.xlsx）
        special_versions_data: 特殊版本数据（d_base、no-explore-clarify、no-clarify），可选
    """
    # 从源文件加载现有工作簿
    if not source_excel_path.exists():
        raise FileNotFoundError(f"源文件不存在: {source_excel_path}")
    
    wb = load_workbook(source_excel_path)
    
    # 检查并创建汇总sheet（各迭代无效需求率平均值）
    # 放在最前面（index=0）
    summary_sheet_name = "Invalid Rate Summary"
    if summary_sheet_name in wb.sheetnames:
        # 如果已存在，删除后重新创建
        wb.remove(wb[summary_sheet_name])
    summary_ws = wb.create_sheet(title=summary_sheet_name, index=0)
    summary_headers = ["Iteration", "Avg Invalid Rate (%)"]
    summary_ws.append(summary_headers)
    
    # 设置汇总表头样式
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = summary_ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 计算并写入各迭代的无效需求率平均值
    for iter_num in sorted(all_data.keys()):
        data = all_data[iter_num]
        invalid_rates = [d["invalid_rate"] for d in data if d["invalid_rate"] is not None]
        if invalid_rates:
            avg_rate = sum(invalid_rates) / len(invalid_rates)
            summary_ws.append([iter_num, round(avg_rate, 2)])
        else:
            summary_ws.append([iter_num, None])
    
    # 设置汇总sheet列宽和对齐
    summary_ws.column_dimensions['A'].width = 15  # 迭代编号
    summary_ws.column_dimensions['B'].width = 20  # 无效需求率平均值
    for row in summary_ws.iter_rows(min_row=2, max_row=summary_ws.max_row):
        for col_idx in [1, 2]:  # A, B列
            cell = row[col_idx - 1]
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 创建权重配置sheet（放在汇总之后）
    create_weight_config_sheet(wb, sheet_index=1)
    
    # 先创建特殊版本的sheet（放在汇总之后，迭代数据之前）
    current_index = 2
    if special_versions_data:
        version_order = ["d_base", "no-explore-clarify", "no-clarify"]
        for version_key in version_order:
            if version_key not in special_versions_data:
                continue
            
            data = special_versions_data[version_key]
            if not data:
                continue
            
            sheet_name = version_key
            # 检查并处理sheet名称冲突
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
            
            ws = wb.create_sheet(title=sheet_name, index=current_index)
            current_index += 1
            
            # 表头
            headers = ["Doc Name", "Pool Size", "Valid Semantic Units Count", "Invalid Rate"]
            ws.append(headers)
            
            # 设置表头样式
            header_font = Font(bold=True, size=12)
            header_alignment = Alignment(horizontal='center', vertical='center')
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.alignment = header_alignment
            
            # 写入数据
            for row_data in data:
                ws.append([
                    row_data["doc_name"],
                    row_data["pool_size"],
                    row_data["valid_semantic_units_count"],
                    round(row_data["invalid_rate"], 2) if row_data["invalid_rate"] is not None else None,
                ])
            
            # 设置列宽
            ws.column_dimensions['A'].width = 50  # 文档名称
            ws.column_dimensions['B'].width = 15  # 需求池数量
            ws.column_dimensions['C'].width = 18  # 有效需求单元数量
            ws.column_dimensions['D'].width = 15  # 无效需求率
            
            # 设置数据对齐
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for col_idx in [2, 3, 4]:  # B, C, D列（数字列）
                    cell = row[col_idx - 1]
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                # A列左对齐
                cell = row[0]
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # 表头（按逻辑分组：基础信息 → 迭代数据 → SRS相关 → 语义单元相关 → 计算指标）
    headers = ["Doc Name", "Check Item Count", "Valid Semantic Units Count", "Pool Size", "SRS Passed Count", "SRS Merge Passed Count", "Units Passed Count", "Units Merge Passed Count", "Invalid Rate"]
    
    # 为每个迭代创建sheet（放在特殊版本之后）
    for iter_num in sorted(all_data.keys()):
        data = all_data[iter_num]
        
        # 创建sheet（sheet名称不能超过31个字符）
        sheet_name = f"iter_{iter_num}"
        # 检查sheet名称冲突，如果已存在则添加后缀
        original_sheet_name = sheet_name
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{original_sheet_name}_{counter}"
            counter += 1
        ws = wb.create_sheet(title=sheet_name, index=current_index)
        current_index += 1
        
        # 设置表头
        ws.append(headers)
        
        # 设置表头样式
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 写入数据
        for row_data in data:
            ws.append([
                row_data["doc_name"],
                row_data["check_item_count"],
                row_data["valid_semantic_units_count"],
                row_data["pool_size"],
                row_data["srs_passed_count"],
                row_data["srs_merge_passed_count"],
                row_data["units_passed_count"],
                row_data["units_merge_passed_count"],
                row_data["invalid_rate"],
            ])
        
        # 设置列宽
        ws.column_dimensions['A'].width = 50  # 文档名称
        ws.column_dimensions['B'].width = 15  # 检查项数量
        ws.column_dimensions['C'].width = 18  # 有效需求单元数量
        ws.column_dimensions['D'].width = 15  # 需求池数量
        ws.column_dimensions['E'].width = 18  # SRS检查项通过数
        ws.column_dimensions['F'].width = 20  # SRS阶段合并通过数
        ws.column_dimensions['G'].width = 20  # 语义单元检查项通过数
        ws.column_dimensions['H'].width = 22  # 语义单元阶段合并通过数
        ws.column_dimensions['I'].width = 15  # 无效需求率
        
        # 设置数据对齐
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for col_idx in [2, 3, 4, 5, 6, 7, 8, 9]:  # B, C, D, E, F, G, H, I列（数字列）
                cell = row[col_idx - 1]
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 创建按通过率分段的汇总sheet（使用iter1_to_last的合并阶段数据）
    # 放在迭代数据之后
    if all_data:
        max_iter = max(all_data.keys())
        create_pass_rate_segmentation_sheet(wb, all_data, eval_reports_dir, output_dir, max_iter, current_index)
        current_index += 1
        
        create_invalid_rate_segmentation_sheet(wb, all_data, eval_reports_dir, output_dir, max_iter, current_index)
    
    wb.save(output_path)
    print(f"Excel文件已保存到: {output_path}，共 {len(all_data)} 个迭代的统计数据")


def main():
    """主函数"""
    # 设置路径
    eval_reports_dir = Path("/root/project/srs/srs-gen2/eval_reports/minimal_en_iter8_merge_passes_1judge_pass_Loose2")
    output_dir = Path("/root/project/srs/srs-gen2/output/minimal_en_iter8/")
    source_excel_path = eval_reports_dir / "数据汇总v2.xlsx"
    output_excel = eval_reports_dir / "数据汇总v3.xlsx"
    
    # 检查源文件是否存在
    if not source_excel_path.exists():
        print(f"错误：源文件不存在: {source_excel_path}")
        return
    
    print("开始收集统计数据...")
    
    # 自动检测所有迭代
    stats_file = output_dir / "all_iter_stats.json"
    available_iters = get_available_iterations(eval_reports_dir, output_dir, stats_file)
    
    if not available_iters:
        print("警告：未找到任何迭代数据")
        return
    
    print(f"检测到以下迭代: {sorted(available_iters)}")
    
    # 收集所有迭代的数据
    all_data = collect_statistics(eval_reports_dir, output_dir, list(available_iters))
    
    # 显示每个迭代的统计摘要
    total_docs = 0
    for iter_num in sorted(all_data.keys()):
        data = all_data[iter_num]
        total_docs += len(data)
        print(f"\n迭代 {iter_num} - 共收集到 {len(data)} 个文档的统计数据")
        if data:
            print(f"  前3条数据预览:")
            for i, row in enumerate(data[:3], 1):
                print(f"    {i}. {row['doc_name']}: 检查项={row['check_item_count']}, 有效需求单元={row['valid_semantic_units_count']}, SRS通过数={row['srs_passed_count']}")
    
    # 收集特殊版本数据（d_base、no-explore-clarify、no-clarify）
    print("\n开始收集特殊版本统计数据...")
    special_versions_data = collect_statistics_for_special_versions(eval_reports_dir, output_dir)
    
    for version_name, data in special_versions_data.items():
        if data:
            print(f"  {version_name}: 共收集到 {len(data)} 个文档的统计数据")
    
    # 保存到Excel
    print(f"\n正在保存到Excel（共 {len(all_data)} 个迭代）...")
    save_to_excel(all_data, output_excel, eval_reports_dir, output_dir, source_excel_path, special_versions_data)
    
    # 显示所有迭代的统计摘要
    print("\n" + "=" * 60)
    print("所有迭代统计摘要:")
    print("=" * 60)
    
    for iter_num in sorted(all_data.keys()):
        data = all_data[iter_num]
        check_counts = [d["check_item_count"] for d in data if d["check_item_count"] is not None]
        unit_counts = [d["valid_semantic_units_count"] for d in data if d["valid_semantic_units_count"] is not None]
        srs_passed_counts = [d["srs_passed_count"] for d in data if d["srs_passed_count"] is not None]
        pool_sizes = [d["pool_size"] for d in data if d["pool_size"] is not None]
        units_passed_counts = [d["units_passed_count"] for d in data if d["units_passed_count"] is not None]
        invalid_rates = [d["invalid_rate"] for d in data if d["invalid_rate"] is not None]
        srs_merge_passed_counts = [d["srs_merge_passed_count"] for d in data if d["srs_merge_passed_count"] is not None]
        units_merge_passed_counts = [d["units_merge_passed_count"] for d in data if d["units_merge_passed_count"] is not None]
        
        print(f"\n迭代 {iter_num}:")
        if check_counts:
            print(f"  检查项数量 - 平均值: {sum(check_counts)/len(check_counts):.2f}, 范围: {min(check_counts)}-{max(check_counts)}")
        if unit_counts:
            print(f"  有效需求单元数量 - 平均值: {sum(unit_counts)/len(unit_counts):.2f}, 范围: {min(unit_counts)}-{max(unit_counts)}")
        if srs_passed_counts:
            print(f"  SRS检查项通过数 - 平均值: {sum(srs_passed_counts)/len(srs_passed_counts):.2f}, 范围: {min(srs_passed_counts)}-{max(srs_passed_counts)}")
        if pool_sizes:
            print(f"  需求池数量 - 平均值: {sum(pool_sizes)/len(pool_sizes):.2f}, 范围: {min(pool_sizes)}-{max(pool_sizes)}")
        if units_passed_counts:
            print(f"  语义单元检查项通过数 - 平均值: {sum(units_passed_counts)/len(units_passed_counts):.2f}, 范围: {min(units_passed_counts)}-{max(units_passed_counts)}")
        if invalid_rates:
            print(f"  无效需求率 - 平均值: {sum(invalid_rates)/len(invalid_rates):.2f}%, 范围: {min(invalid_rates):.2f}%-{max(invalid_rates):.2f}%")
        if srs_merge_passed_counts:
            print(f"  SRS阶段合并通过数 - 平均值: {sum(srs_merge_passed_counts)/len(srs_merge_passed_counts):.2f}, 范围: {min(srs_merge_passed_counts)}-{max(srs_merge_passed_counts)}")
        if units_merge_passed_counts:
            print(f"  语义单元阶段合并通过数 - 平均值: {sum(units_merge_passed_counts)/len(units_merge_passed_counts):.2f}, 范围: {min(units_merge_passed_counts)}-{max(units_merge_passed_counts)}")


if __name__ == "__main__":
    main()

