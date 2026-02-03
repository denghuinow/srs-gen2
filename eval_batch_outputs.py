#!/usr/bin/env python3
"""对 srs-gen2/batch_run.py 输出结果批量执行评估"""
import argparse
import csv
import json
import re
import shlex
import shutil
import subprocess
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple


PROJECT_ROOT = Path(__file__).resolve().parent

# 添加 srs-eval 到路径以便导入
SRS_EVAL_DIR = PROJECT_ROOT.parent / "srs-eval"
if SRS_EVAL_DIR.exists():
    sys.path.insert(0, str(SRS_EVAL_DIR))
    try:
        from src.output_formatter import OutputFormatter
        from src.evaluator import DocumentEvaluation, CheckpointResult
    except ImportError:
        OutputFormatter = None
        DocumentEvaluation = None
        CheckpointResult = None
else:
    OutputFormatter = None
    DocumentEvaluation = None
    CheckpointResult = None


@dataclass
class Stage:
    """阶段元数据"""

    name: str
    target_dir: Path
    output_dir: Path
    file_count: int


@dataclass
class DocumentScore:
    """单个文档的评估指标"""

    weighted: Optional[float]
    voting: Optional[float]
    average: Optional[float]
    categories: Dict[str, Optional[float]]


@dataclass
class StageEvalSummary:
    """单个阶段的评估结果集合"""

    raw_name: str
    display_name: str
    documents: Dict[str, DocumentScore]


REFERENCE_STAGE_NAME = "r_base"


DIMENSION_ORDER = [
    "FUNCTIONAL",
    "BUSINESS_FLOW",
    "BOUNDARY",
    "EXCEPTION",
    "DATA_STATE",
    "CONSISTENCY_RULE",
    "SECURITY",
]


def safe_mean(values: Iterable[Optional[float]]) -> Optional[float]:
    """计算平均值，自动跳过 None"""

    materialized = [v for v in values if v is not None]
    if not materialized:
        return None
    return sum(materialized) / len(materialized)


def format_score(value: Optional[float], digits: int = 2) -> Optional[float]:
    """保留两位小数（或指定精度）"""

    if value is None:
        return None
    return round(value, digits)


def format_ratio(value: Optional[float]) -> Optional[float]:
    """用于通过率的格式化"""

    if value is None:
        return None
    return round(value, 4)


def normalize_doc_stage_name(raw: str) -> str:
    """文档评估阶段命名"""

    if raw == "srs_baseline":
        return "d_base"
    name = raw
    if name.startswith("srs_"):
        name = name[4:]
    if name.startswith("iter_"):
        suffix = name.split("_", 1)[1]
        return f"iter{suffix}"
    return name


def normalize_unit_stage_name(raw: str) -> str:
    """语义单元评估阶段命名（加 units_ 前缀）"""

    if raw.startswith("pool_iter_"):
        suffix = raw.split("_", 2)[-1]
        return f"units_iter{suffix}"
    if raw in {"no-clarify", "no-explore-clarify"}:
        return f"units_{raw}"
    return f"units_{raw}"


def stage_sort_key(display_name: str) -> Tuple[int, object]:
    """统一的阶段排序规则"""

    name = display_name.lower()
    if name.startswith("req_"):
        name = name[4:]
    if name.startswith("units_"):
        name = name[6:]
    priority = {
        "r_base": 0,
        "d_base": 1,
        "baseline": 1,
        "no-explore-clarify": 2,
        "no-clarify": 3,
    }
    if name in priority:
        return (priority[name], 0)
    if name.startswith("iter"):
        suffix = name[4:]
        # 处理合并阶段名称，如 iter_1_to_2
        if "_to_" in suffix:
            # 提取最大迭代编号用于排序
            try:
                max_iter = int(suffix.split("_to_")[-1])
                return (5, max_iter)  # 合并阶段排在普通迭代阶段之后
            except (ValueError, IndexError):
                return (5, suffix)
        else:
            # 普通迭代阶段
            try:
                return (4, int(suffix))
            except ValueError:
                return (4, suffix)
    return (6, name)


def collect_document_names(stages: Sequence[StageEvalSummary]) -> List[str]:
    """获取所有文档名称，按字母排序"""

    names = {doc for stage in stages for doc in stage.documents}
    return sorted(names)


def load_stage_evaluations(stage_root: Path, name_normalizer) -> List[StageEvalSummary]:
    """从评估输出目录加载阶段数据"""

    if not stage_root.exists():
        return []
    stages: List[StageEvalSummary] = []
    for directory in sorted(stage_root.iterdir(), key=lambda p: p.name.lower()):
        if not directory.is_dir():
            continue
        documents: Dict[str, DocumentScore] = {}
        for json_file in sorted(directory.glob("*_evaluation.json")):
            try:
                data = json.loads(json_file.read_text(encoding="utf-8"))
            except json.JSONDecodeError as exc:
                print(f"⚠ 无法解析 {json_file}: {exc}")
                continue
            doc_name = json_file.name.replace("_evaluation.json", "")
            scores = data.get("scores", {})
            category_scores = {
                cat: info.get("score")
                for cat, info in (scores.get("categories") or {}).items()
            }
            documents[doc_name] = DocumentScore(
                weighted=scores.get("weighted_score"),
                voting=scores.get("voting_score"),
                average=scores.get("average_score"),
                categories=category_scores,
            )
        if documents:
            stages.append(
                StageEvalSummary(
                    raw_name=directory.name,
                    display_name=name_normalizer(directory.name),
                    documents=documents,
                )
            )
    stages.sort(key=lambda stage: stage_sort_key(stage.display_name))
    return stages


def build_stage_score_rows(
    stages: Sequence[StageEvalSummary],
    include_reference_row: bool = True,
) -> Tuple[List[str], List[List[Optional[float]]]]:
    """生成逐阶段-逐文档的分数矩阵"""

    doc_names = collect_document_names(stages)
    rows: List[List[Optional[float]]] = []
    reference_stage = next((stage for stage in stages if stage.display_name == REFERENCE_STAGE_NAME), None)

    if include_reference_row:
        summary_row: List[Optional[float]] = [0, REFERENCE_STAGE_NAME]
        if doc_names:
            if reference_stage:
                for doc in doc_names:
                    doc_score = reference_stage.documents.get(doc)
                    summary_row.append(format_score(doc_score.weighted if doc_score else None))
            else:
                summary_row.extend(["-"] * len(doc_names))
        rows.append(summary_row)

    seq_counter = 0
    for stage in stages:
        if stage is reference_stage:
            continue
        seq_counter += 1
        row = [seq_counter, stage.display_name]
        for doc in doc_names:
            doc_score = stage.documents.get(doc)
            row.append(format_score(doc_score.weighted if doc_score else None))
        rows.append(row)
    return doc_names, rows


def build_stage_summary_rows(
    stages: Sequence[StageEvalSummary],
    include_reference_row: bool = True,
) -> List[List[Optional[float]]]:
    """阶段整体统计"""

    rows: List[List[Optional[float]]] = []
    reference_stage = next((stage for stage in stages if stage.display_name == REFERENCE_STAGE_NAME), None)
    if include_reference_row:
        if reference_stage:
            doc_scores = list(reference_stage.documents.values())
            rows.append(
                [
                    0,
                    REFERENCE_STAGE_NAME,
                    len(doc_scores),
                    format_score(safe_mean(score.weighted for score in doc_scores)),
                    format_ratio(
                        safe_mean(
                            (score.voting / 100.0)
                            for score in doc_scores
                            if score.voting is not None
                        )
                    ),
                    format_ratio(
                        safe_mean(
                            (score.average / 100.0)
                            for score in doc_scores
                            if score.average is not None
                        )
                    ),
                ]
            )
        else:
            rows.append([0, REFERENCE_STAGE_NAME, "-", "-", "-", "-"])

    seq_counter = 0
    for stage in stages:
        if stage is reference_stage:
            continue
        doc_scores = list(stage.documents.values())
        if not doc_scores:
            continue
        seq_counter += 1
        rows.append(
            [
                seq_counter,
                stage.display_name,
                len(doc_scores),
                format_score(safe_mean(score.weighted for score in doc_scores)),
                format_ratio(
                    safe_mean(
                        (score.voting / 100.0)
                        for score in doc_scores
                        if score.voting is not None
                    )
                ),
                format_ratio(
                    safe_mean(
                        (score.average / 100.0)
                        for score in doc_scores
                        if score.average is not None
                    )
                ),
            ]
        )
    return rows


def _dimension_values(stage_subset: Sequence[StageEvalSummary]) -> List[Optional[float]]:
    doc_scores = [score for stage in stage_subset for score in stage.documents.values()]
    values: List[Optional[float]] = []
    for dim in DIMENSION_ORDER:
        dim_scores = [
            score.categories.get(dim)
            for score in doc_scores
            if score.categories.get(dim) is not None
        ]
        values.append(format_ratio(safe_mean(val / 100.0 for val in dim_scores)))
    values.append(
        format_ratio(
            safe_mean(
                (score.voting / 100.0)
                for score in doc_scores
                if score.voting is not None
            )
        )
    )
    values.append(
        format_ratio(
            safe_mean(
                (score.average / 100.0)
                for score in doc_scores
                if score.average is not None
            )
        )
    )
    return values


def build_dimension_rows(
    stages: Sequence[StageEvalSummary],
    include_reference_row: bool = True,
) -> List[List[Optional[float]]]:
    """每个阶段的维度统计"""

    rows: List[List[Optional[float]]] = []
    reference_stage = next((stage for stage in stages if stage.display_name == REFERENCE_STAGE_NAME), None)
    if include_reference_row:
        if reference_stage:
            rows.append([0, REFERENCE_STAGE_NAME, *_dimension_values([reference_stage])])
        else:
            rows.append([0, REFERENCE_STAGE_NAME, *(["-"] * (len(DIMENSION_ORDER) + 2))])

    seq_counter = 0
    for stage in stages:
        if stage is reference_stage:
            continue
        seq_counter += 1
        rows.append([seq_counter, stage.display_name, *_dimension_values([stage])])
    return rows


def write_stage_score_sheet(
    ws,
    stages: Sequence[StageEvalSummary],
    column_label: str,
    include_reference_row: bool = True,
) -> None:
    doc_names, rows = build_stage_score_rows(stages, include_reference_row)
    header = ["Seq", column_label]
    if doc_names:
        header.extend(doc_names)
    ws.append(header)
    if not rows:
        ws.append(["-", "No evaluation results available"])
        return
    for row in rows:
        ws.append(row)


def write_stage_summary_sheet(
    ws,
    stages: Sequence[StageEvalSummary],
    include_reference_row: bool = True,
) -> None:
    ws.append(["Seq", "Stage", "Doc Count", "Avg Weighted Score", "Avg Voting Pass Rate", "Avg Pass Rate"])
    rows = build_stage_summary_rows(stages, include_reference_row)
    if not rows:
        ws.append(["-", "No evaluation stages available"])
        return
    for row in rows:
        ws.append(row)


def write_dimension_sheet(
    ws,
    stages: Sequence[StageEvalSummary],
    include_reference_row: bool = True,
) -> None:
    ws.append(["Seq", "Stage", *DIMENSION_ORDER, "Avg Voting Pass Rate", "Avg Pass Rate"])
    rows = build_dimension_rows(stages, include_reference_row)
    if not rows:
        ws.append(["-", "No dimension data available"])
        return
    for row in rows:
        ws.append(row)


def load_time_stats(csv_path: Path) -> Optional[List[List[str]]]:
    """读取耗时统计 CSV"""

    if not csv_path.exists():
        return None
    with csv_path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        return [row for row in reader]


def write_time_sheet(ws, time_rows: Optional[List[List[str]]], csv_path: Path) -> None:
    if not time_rows:
        ws.append(["Note", f"Time statistics file not found: {csv_path}"])
        return
    for row in time_rows:
        ws.append(row)


def extract_iter_number(stage_name: str) -> Optional[int]:
    """从阶段名称中提取迭代编号
    
    Args:
        stage_name: 阶段名称，如 "iter1", "iter2", "units_iter3"
    
    Returns:
        迭代编号，如果无法提取则返回 None
    """
    # 匹配 iter 后面的数字
    match = re.search(r'iter(\d+)', stage_name, re.IGNORECASE)
    if match:
        return int(match.group(1))
    return None


def merge_iter_evaluations(
    stage_root: Path,
    name_normalizer,
    is_units: bool = False,
) -> None:
    """合并 iter_1 到 iter_n 的评估结果
    
    Args:
        stage_root: 阶段根目录
        name_normalizer: 阶段名称规范化函数
        is_units: 是否为语义单元阶段
    """
    if OutputFormatter is None or DocumentEvaluation is None or CheckpointResult is None:
        print("⚠ 无法导入 srs-eval 模块，跳过合并迭代通过项功能")
        return
    
    if not stage_root.exists():
        return
    
    # 加载所有阶段（排除合并阶段）
    all_stages = []
    for directory in sorted(stage_root.iterdir(), key=lambda p: p.name.lower()):
        if not directory.is_dir():
            continue
        # 跳过合并阶段（包含 _to_ 的目录名）
        if "_to_" in directory.name:
            continue
        display_name = name_normalizer(directory.name)
        iter_num = extract_iter_number(display_name)
        if iter_num is not None:
            all_stages.append((iter_num, display_name, directory))
    
    if not all_stages:
        return
    
    # 按迭代编号排序
    all_stages.sort(key=lambda x: x[0])
    
    # 提取实际存在的迭代编号集合（去重并排序）
    existing_iter_nums = sorted(set(num for num, _, _ in all_stages))
    
    if not existing_iter_nums:
        return
    
    # 找到所有文档名称
    all_doc_names = set()
    for _, _, directory in all_stages:
        for json_file in directory.glob("*_evaluation.json"):
            doc_name = json_file.name.replace("_evaluation.json", "")
            all_doc_names.add(doc_name)
    
    if not all_doc_names:
        return
    
    # 为每个实际存在的迭代编号创建合并阶段
    for max_iter in existing_iter_nums:
        # 找到 iter_1 到 iter_max_iter 的所有阶段
        iter_stages = [(num, name, dir_path) for num, name, dir_path in all_stages if num <= max_iter]
        
        if len(iter_stages) <= 1:
            # 只有一个阶段，不需要合并
            continue
        
        # 创建合并阶段名称（使用原始阶段名称格式，而不是规范化后的名称）
        if is_units:
            # 语义单元阶段：pool_iter_1_to_n
            merged_stage_name = f"pool_iter_1_to_{max_iter}"
        else:
            # 文档阶段：srs_iter_1_to_n
            merged_stage_name = f"srs_iter_1_to_{max_iter}"
        merged_stage_dir = stage_root / merged_stage_name
        merged_stage_dir.mkdir(parents=True, exist_ok=True)
        
        print(f"合并阶段 {merged_stage_name}（包含 iter_1 到 iter_{max_iter}）")
        
        # 对每个文档进行合并
        for doc_name in sorted(all_doc_names):
            evaluations = []
            checkpoints = None
            
            # 加载所有相关阶段的评估结果
            for iter_num, stage_name, stage_dir in iter_stages:
                json_file = stage_dir / f"{doc_name}_evaluation.json"
                if not json_file.exists():
                    continue
                
                evaluation = OutputFormatter.load_json(json_file)
                if evaluation is None:
                    continue
                
                evaluations.append(evaluation)
                
                # 使用第一个阶段的检查项列表作为基准
                if checkpoints is None:
                    checkpoints = evaluation.checkpoints
            
            if not evaluations:
                continue
            
            if checkpoints is None:
                continue
            
            # 合并检查项结果（OR 逻辑：任意一个阶段通过就算通过）
            # 假设所有阶段使用相同的检查项列表（从同一个基准文档提取）
            merged_checkpoint_results = []
            for cp_idx, checkpoint in enumerate(checkpoints):
                passed_in_any = False
                passed_count = 0
                total_stages = len(evaluations)
                
                # 从所有阶段收集该检查项的结果
                category = None
                checkpoint_content = None
                raw_checkpoint = checkpoint
                
                for evaluation in evaluations:
                    # 确保索引在范围内
                    if cp_idx < len(evaluation.checkpoint_results):
                        cp_result = evaluation.checkpoint_results[cp_idx]
                        if cp_result.passed:
                            passed_in_any = True
                            passed_count += 1
                        # 从第一个阶段获取元信息
                        if category is None:
                            category = cp_result.category
                            checkpoint_content = cp_result.checkpoint
                            raw_checkpoint = cp_result.raw_checkpoint
                    elif cp_idx < len(evaluation.checkpoints):
                        # 如果检查项结果数量不匹配，但检查项列表中有，尝试匹配
                        # 这种情况不应该发生，但为了健壮性处理
                        pass
                
                # 计算 pass_rate（通过的阶段数 / 总阶段数）
                pass_rate = passed_count / total_stages if total_stages > 0 else 0.0
                
                # 如果没有找到检查项内容，从 checkpoint 字符串中提取
                if checkpoint_content is None:
                    checkpoint_content = checkpoint.split("] ", 1)[-1] if "] " in checkpoint else checkpoint
                
                merged_cp_result = CheckpointResult(
                    checkpoint=checkpoint_content,
                    passed=passed_in_any,
                    pass_rate=pass_rate,
                    category=category,
                    raw_checkpoint=raw_checkpoint,
                )
                merged_checkpoint_results.append(merged_cp_result)
            
            # 创建合并后的评估结果
            merged_evaluation = DocumentEvaluation(
                target_document=evaluations[0].target_document,
                checkpoints=checkpoints,
                checkpoint_results=merged_checkpoint_results,
                all_judge_results=None,  # 合并阶段不保留多评委结果
                model_name=evaluations[0].model_name,
                baseline_document=evaluations[0].baseline_document,
            )
            
            # 保存合并后的评估结果
            merged_json_file = merged_stage_dir / f"{doc_name}_evaluation.json"
            merged_md_file = merged_stage_dir / f"{doc_name}_evaluation.md"
            merged_tsv_file = merged_stage_dir / f"{doc_name}_evaluation.tsv"
            
            # 保存 JSON
            json_data = OutputFormatter.to_json(merged_evaluation)
            merged_json_file.write_text(json.dumps(json_data, ensure_ascii=False, indent=2), encoding="utf-8")
            
            # 保存 Markdown
            md_content = OutputFormatter.to_markdown(merged_evaluation)
            merged_md_file.write_text(md_content, encoding="utf-8")
            
            # 保存 TSV
            tsv_content = OutputFormatter.to_csv([merged_evaluation])
            merged_tsv_file.write_text(tsv_content, encoding="utf-8")
        
        print(f"✓ 合并阶段 {merged_stage_name} 完成")


def generate_summary_workbook(
    summary_path: Path,
    doc_eval_root: Path,
    unit_eval_root: Path,
    outputs_dir: Path,
) -> None:
    """生成类似数据汇总v2.xlsx 的统计报表"""

    try:
        from openpyxl import Workbook
    except ImportError:
        print("⚠ 未安装 openpyxl，跳过汇总 Excel 生成")
        return

    doc_stages = load_stage_evaluations(doc_eval_root, normalize_doc_stage_name)
    unit_stages = load_stage_evaluations(unit_eval_root, normalize_unit_stage_name)
    summary_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    doc_sheet = wb.active
    doc_sheet.title = "Document by Stage"
    write_stage_score_sheet(doc_sheet, doc_stages, "Stage/Document")

    doc_summary_sheet = wb.create_sheet("Document Stage Summary")
    write_stage_summary_sheet(doc_summary_sheet, doc_stages)

    doc_dimension_sheet = wb.create_sheet("Document Dimensions")
    write_dimension_sheet(doc_dimension_sheet, doc_stages)

    unit_sheet = wb.create_sheet("Semantic Unit by Stage")
    write_stage_score_sheet(unit_sheet, unit_stages, "Stage", include_reference_row=False)

    unit_summary_sheet = wb.create_sheet("Semantic Unit Stage Summary")
    write_stage_summary_sheet(unit_summary_sheet, unit_stages, include_reference_row=False)

    unit_dimension_sheet = wb.create_sheet("Semantic Unit Dimensions")
    write_dimension_sheet(unit_dimension_sheet, unit_stages, include_reference_row=False)

    time_sheet = wb.create_sheet("Generation Time")
    time_rows = load_time_stats(outputs_dir / "耗时统计-总耗时.csv")
    write_time_sheet(time_sheet, time_rows, outputs_dir / "耗时统计-总耗时.csv")

    wb.save(summary_path)
    print(f"汇总 Excel 已写入：{summary_path}")


def parse_extra_args(extra: Optional[str]) -> List[str]:
    """解析传递给评估脚本的额外参数"""
    if not extra:
        return []
    return shlex.split(extra)


def list_stage_dirs(root: Path) -> List[Path]:
    """列出阶段目录"""
    if not root.exists():
        return []
    return sorted([p for p in root.iterdir() if p.is_dir()], key=lambda p: p.name.lower())


def prepare_doc_stage(stage_dir: Path, temp_root: Path) -> tuple[Optional[Path], List[str], bool]:
    """准备 SRS 文档阶段，必要时把 .txt 转存为 .md"""

    markdown_files = sorted(
        [file for file in stage_dir.iterdir() if file.is_file() and file.suffix.lower() in {".md", ".markdown"}],
        key=lambda p: p.name.lower(),
    )
    if markdown_files:
        return stage_dir, [file.stem for file in markdown_files], False

    txt_files = sorted(
        [file for file in stage_dir.iterdir() if file.is_file() and file.suffix.lower() == ".txt"],
        key=lambda p: p.name.lower(),
    )
    if not txt_files:
        return None, [], False

    temp_root.mkdir(parents=True, exist_ok=True)
    temp_dir = temp_root / stage_dir.name
    if temp_dir.exists():
        shutil.rmtree(temp_dir)
    temp_dir.mkdir(parents=True, exist_ok=True)

    doc_names: List[str] = []
    for file in txt_files:
        target_path = temp_dir / f"{file.stem}.md"
        target_path.write_text(file.read_text(encoding="utf-8"), encoding="utf-8")
        doc_names.append(file.stem)

    return temp_dir, doc_names, True


def prepare_unit_stage(stage_dir: Path, temp_root: Path) -> tuple[Path, int]:
    """将语义单元（txt/md）转存为 md，方便主评估脚本加载"""
    dest = temp_root / stage_dir.name
    if dest.exists():
        shutil.rmtree(dest)
    dest.mkdir(parents=True, exist_ok=True)
    count = 0
    for file in sorted(stage_dir.iterdir(), key=lambda p: p.name.lower()):
        if not file.is_file():
            continue
        suffix = file.suffix.lower()
        if suffix not in {".md", ".markdown", ".txt"}:
            continue
        target_path = dest / f"{file.stem}.md"
        target_path.write_text(file.read_text(encoding="utf-8"), encoding="utf-8")
        count += 1
    return dest, count


def prepare_r_base_stage(
    r_base_dir: Path,
    doc_names: Optional[Sequence[str]],
    temp_root: Path,
) -> tuple[Optional[Path], int]:
    """将 --r-base-dir 中的文档复制到临时 r_base 阶段目录"""

    if not r_base_dir.exists():
        print(f"⚠ r_base 阶段：r-base 目录不存在：{r_base_dir}")
        return None, 0

    stage_dir = temp_root / "r_base"
    if stage_dir.exists():
        shutil.rmtree(stage_dir)
    stage_dir.mkdir(parents=True, exist_ok=True)

    allowed_suffixes = {".md", ".markdown", ".txt"}
    source_map = {
        file.stem.lower(): file
        for file in r_base_dir.iterdir()
        if file.is_file() and file.suffix.lower() in allowed_suffixes
    }
    if not source_map:
        print(f"⚠ r_base 阶段：r-base 目录中没有可用的 .md/.markdown/.txt 文件：{r_base_dir}")
        shutil.rmtree(stage_dir, ignore_errors=True)
        return None, 0

    if doc_names:
        ordered_names = sorted(dict.fromkeys(doc_names), key=lambda name: name.lower())
    else:
        ordered_names = sorted(
            {file.stem for file in source_map.values()},
            key=lambda name: name.lower(),
        )

    missing: List[str] = []
    copied = 0
    for name in ordered_names:
        source = source_map.get(name.lower())
        if not source:
            missing.append(name)
            continue
        dest_path = stage_dir / f"{name}.md"
        shutil.copyfile(source, dest_path)
        copied += 1

    if missing:
        print(f"⚠ r_base 阶段：以下文档未在 r-base 目录找到：{', '.join(missing)}")

    if copied == 0:
        shutil.rmtree(stage_dir, ignore_errors=True)
        return None, 0

    return stage_dir, copied


def build_eval_cmd(
    runner: str,
    python_bin: str,
    uv_bin: str,
    srs_eval_main: Path,
    srs_eval_dir: Path,
    baseline_dir: Path,
    target_dir: Path,
    output_dir: Path,
    extra_args: Sequence[str],
) -> tuple[List[str], Optional[Path]]:
    """构建调用 srs-eval/main.py 的命令"""
    if runner == "uv":
        cmd = [
            uv_bin,
            "run",
            "--project",
            str(srs_eval_dir),
            str(srs_eval_main),
        ]
        workdir: Optional[Path] = srs_eval_dir
    else:
        cmd = [
            python_bin,
            str(srs_eval_main),
        ]
        workdir = None

    cmd.extend(
        [
            "--baseline-dir",
            str(baseline_dir),
            "--target-dir",
            str(target_dir),
            "--output-dir",
            str(output_dir),
        ]
    )
    cmd.extend(extra_args)
    return cmd, workdir


def run_stage(stage: Stage, cmd: Sequence[str], dry_run: bool, workdir: Optional[Path]) -> tuple[Stage, int]:
    """执行单阶段评估"""
    print(f"\n[阶段 {stage.name}] 待评估文件：{stage.file_count}")
    print(f"阶段输出目录：{stage.output_dir}")
    print("执行命令：", " ".join(shlex.quote(p) for p in cmd))
    if dry_run:
        return stage, 0
    proc = subprocess.run(cmd, cwd=str(workdir) if workdir else None)
    return stage, proc.returncode


StageJob = Tuple[Stage, Sequence[str], bool, Optional[Path]]


def summarize(
    results: Iterable[tuple[Stage, int]],
    label: str,
    skip_stage_names: Optional[Sequence[str]] = None,
) -> None:
    """打印汇总信息"""

    results = list(results)
    skip_names = set(skip_stage_names or [])
    if not results:
        print(f"\n[{label}] 未检测到任何阶段")
        return

    filtered = [item for item in results if item[0].name not in skip_names]
    if not filtered:
        print(f"\n[{label}] 未检测到任何阶段")
        if skip_names:
            print(f"{label} 提示：已过滤阶段：{', '.join(sorted(skip_names))}")
        return

    success = [stage for stage, code in filtered if code == 0]
    failed = [(stage, code) for stage, code in filtered if code != 0]
    print(f"\n[{label}] 阶段总数：{len(filtered)}")
    print(f"{label} 成功：{len(success)}")
    if failed:
        print(f"{label} 失败：{len(failed)}")
        for stage, code in failed:
            print(f"  - {stage.name}: 退出码 {code}")
    else:
        print(f"{label} 全部成功 ✅")

    if skip_names:
        skipped_failed = [
            (stage, code) for stage, code in results if stage.name in skip_names and code != 0
        ]
        if skipped_failed:
            print(f"{label}（未计入统计）阶段失败：")
            for stage, code in skipped_failed:
                print(f"  - {stage.name}: 退出码 {code}")


def main() -> None:
    parser = argparse.ArgumentParser(description="批量评估 srs-gen2/batch_run.py 输出的辅助脚本")
    parser.add_argument(
        "--outputs-dir",
        default=str(PROJECT_ROOT / "output"),
        help="batch_run.py 的输出根目录（默认：项目内 output）",
    )
    parser.add_argument(
        "--d-orig-dir",
        required=True,
        help="原始 d-orig SRS 文档目录，作为评估基准",
    )
    parser.add_argument(
        "--units-baseline-dir",
        help="语义单元评估使用的基准目录（默认与 --d-orig-dir 相同）",
    )
    parser.add_argument(
        "--r-base-dir",
        help="batch_run.py 的 --r-base-dir 目录，用于 summary 阶段（默认：<outputs-dir>/srs_collection/r_base）",
    )
    parser.add_argument(
        "--srs-eval-dir",
        default=str(PROJECT_ROOT.parent / "srs-eval"),
        help="srs-eval 项目路径（默认：与 srs-gen2 平级的 srs-eval）",
    )
    parser.add_argument(
        "--mode",
        choices=["all", "docs", "units"],
        default="all",
        help="选择评估 SRS 文档、语义单元或全部",
    )
    parser.add_argument(
        "--eval-output-dir",
        help="评估结果根目录（默认：<outputs-dir>/eval_reports）",
    )
    parser.add_argument(
        "--summary-xlsx",
        help="生成评估汇总报表的输出路径（默认：<eval-output-dir>/数据汇总v2.xlsx）",
    )
    parser.add_argument(
        "--skip-summary-xlsx",
        action="store_true",
        help="跳过生成 Excel 汇总报表",
    )
    parser.add_argument(
        "--python-bin",
        default=sys.executable,
        help="运行 srs-eval/main.py 的 Python 解释器（默认：当前解释器）",
    )
    parser.add_argument(
        "--srs-eval-runner",
        choices=["python", "uv"],
        default="python",
        help="调用 srs-eval 时使用的运行方式：直接 python 或 uv run（默认：python）",
    )
    parser.add_argument(
        "--uv-bin",
        default="uv",
        help="uv 可执行文件路径（仅在 --srs-eval-runner=uv 时使用）",
    )
    parser.add_argument(
        "--eval-extra-args",
        help='传递给 srs-eval/main.py 的额外参数，例如 "--judges 1 --skip-existing"',
    )
    parser.add_argument(
        "--min-judges-pass",
        type=int,
        default=None,
        help="判定通过的评委数要求（绝对数量）。例如：3个评委中需要2个通过才算通过。默认使用多数投票（超过50%）",
    )
    parser.add_argument(
        "--max-parallel",
        type=int,
        default=1,
        help="同时执行的阶段数量（默认：1，串行）",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="仅打印命令，不实际执行",
    )
    parser.add_argument(
        "--no-cache",
        action="store_true",
        help="强制重新调用模型并覆盖评估缓存（不使用已缓存的评估结果）",
    )
    parser.add_argument(
        "--merge-iter-passes",
        action="store_true",
        help="合并 iter_1 到 iter_n 的通过项（OR 逻辑：任意一个阶段通过就算通过），创建新的合并阶段（如 iter_1_to_n）",
    )
    args = parser.parse_args()

    outputs_dir = Path(args.outputs_dir).expanduser().resolve()
    d_orig_dir = Path(args.d_orig_dir).expanduser().resolve()
    units_baseline_dir = (
        Path(args.units_baseline_dir).expanduser().resolve()
        if args.units_baseline_dir
        else d_orig_dir
    )
    default_r_base_dir = outputs_dir / "srs_collection" / "r_base"
    if args.r_base_dir:
        requirement_base_dir = Path(args.r_base_dir).expanduser().resolve()
    elif default_r_base_dir.exists():
        requirement_base_dir = default_r_base_dir
    else:
        requirement_base_dir = None
    srs_eval_dir = Path(args.srs_eval_dir).expanduser().resolve()
    eval_output_dir = (
        Path(args.eval_output_dir).expanduser().resolve()
        if args.eval_output_dir
        else outputs_dir / "eval_reports"
    )
    eval_output_dir.mkdir(parents=True, exist_ok=True)

    doc_eval_root = eval_output_dir / "documents"
    unit_eval_root = eval_output_dir / "units"
    doc_temp_root = eval_output_dir / ".tmp_doc_inputs"
    units_temp_root = eval_output_dir / ".tmp_units"

    srs_eval_main = srs_eval_dir / "main.py"
    if not srs_eval_main.exists():
        raise SystemExit(f"未找到 srs-eval 主程序：{srs_eval_main}")
    if not outputs_dir.exists():
        raise SystemExit(f"输出目录不存在：{outputs_dir}")
    if not d_orig_dir.exists():
        raise SystemExit(f"d-orig 目录不存在：{d_orig_dir}")
    if not units_baseline_dir.exists():
        raise SystemExit(f"units baseline 目录不存在：{units_baseline_dir}")
    if requirement_base_dir and not requirement_base_dir.exists():
        raise SystemExit(f"r-base 目录不存在：{requirement_base_dir}")

    if args.eval_extra_args:
        extra_args = parse_extra_args(args.eval_extra_args)
    else:
        extra_args = ["--skip-existing"]
    
    # 如果指定了 --no-cache，添加到额外参数中
    if args.no_cache:
        extra_args.append("--no-cache")
    
    # 如果指定了 --min-judges-pass，添加到额外参数中
    if args.min_judges_pass is not None:
        extra_args.extend(["--min-judges-pass", str(args.min_judges_pass)])

    doc_jobs: List[StageJob] = []
    unit_jobs: List[StageJob] = []
    collected_doc_names: set[str] = set()
    r_base_temp_root = eval_output_dir / ".tmp_r_base_inputs"
    cleanup_r_base_temp = False
    cleanup_doc_temp = False

    # 评估 SRS 文档
    if args.mode in ("all", "docs"):
        srs_collection = outputs_dir / "srs_collection"
        doc_stages = list_stage_dirs(srs_collection)
        if not doc_stages:
            print(f"⚠ 未在 {srs_collection} 找到阶段，跳过 SRS 文档评估")
        else:
            doc_eval_root.mkdir(parents=True, exist_ok=True)
            for stage_dir in doc_stages:
                target_dir, doc_names, used_temp = prepare_doc_stage(stage_dir, doc_temp_root)
                if not target_dir or not doc_names:
                    print(f"  - 阶段 {stage_dir.name} 无可评估的 .md/.markdown/.txt 文件，跳过")
                    continue
                if used_temp:
                    cleanup_doc_temp = True
                collected_doc_names.update(doc_names)
                stage_output = doc_eval_root / stage_dir.name
                stage_output.mkdir(parents=True, exist_ok=True)
                stage = Stage(stage_dir.name, target_dir, stage_output, len(doc_names))
                cmd, workdir = build_eval_cmd(
                    args.srs_eval_runner,
                    args.python_bin,
                    args.uv_bin,
                    srs_eval_main,
                    srs_eval_dir,
                    d_orig_dir,
                    target_dir,
                    stage_output,
                    extra_args,
                )
                doc_jobs.append((stage, cmd, args.dry_run, workdir))

    # 准备 r_base 阶段（需求基础文档）
    if requirement_base_dir:
        doc_eval_root.mkdir(parents=True, exist_ok=True)
        target_names = sorted(collected_doc_names) if collected_doc_names else None
        r_base_stage_dir, r_base_count = prepare_r_base_stage(
            requirement_base_dir, target_names, r_base_temp_root
        )
        if r_base_stage_dir and r_base_count:
            stage_output = doc_eval_root / r_base_stage_dir.name
            stage_output.mkdir(parents=True, exist_ok=True)
            stage = Stage(r_base_stage_dir.name, r_base_stage_dir, stage_output, r_base_count)
            cmd, workdir = build_eval_cmd(
                args.srs_eval_runner,
                args.python_bin,
                args.uv_bin,
                srs_eval_main,
                srs_eval_dir,
                d_orig_dir,
                r_base_stage_dir,
                stage_output,
                extra_args,
            )
            doc_jobs.insert(0, (stage, cmd, args.dry_run, workdir))
            cleanup_r_base_temp = not args.dry_run
        else:
            shutil.rmtree(r_base_temp_root, ignore_errors=True)

    # 评估语义单元
    cleanup_units_temp = False

    if args.mode in ("all", "units"):
        units_collection = outputs_dir / "units_collection"
        unit_stages = list_stage_dirs(units_collection)
        if not unit_stages:
            print(f"⚠ 未在 {units_collection} 找到阶段，跳过语义单元评估")
        else:
            units_temp_root.mkdir(parents=True, exist_ok=True)
            unit_eval_root.mkdir(parents=True, exist_ok=True)
            cleanup_units_temp = not args.dry_run
            for stage_dir in unit_stages:
                temp_dir, count = prepare_unit_stage(stage_dir, units_temp_root)
                if count == 0:
                    print(f"  - 阶段 {stage_dir.name} 无可评估文件，跳过")
                    continue
                stage_output = unit_eval_root / stage_dir.name
                stage_output.mkdir(parents=True, exist_ok=True)
                stage = Stage(stage_dir.name, temp_dir, stage_output, count)
                cmd, workdir = build_eval_cmd(
                    args.srs_eval_runner,
                    args.python_bin,
                    args.uv_bin,
                    srs_eval_main,
                    srs_eval_dir,
                    units_baseline_dir,
                    temp_dir,
                    stage_output,
                    extra_args,
                )
                unit_jobs.append((stage, cmd, args.dry_run, workdir))

    doc_results = execute_jobs(doc_jobs, args.max_parallel)
    unit_results = execute_jobs(unit_jobs, args.max_parallel)

    if cleanup_doc_temp and doc_temp_root.exists():
        shutil.rmtree(doc_temp_root, ignore_errors=True)
    if cleanup_units_temp and units_temp_root.exists():
        shutil.rmtree(units_temp_root, ignore_errors=True)
    if cleanup_r_base_temp and r_base_temp_root.exists():
        shutil.rmtree(r_base_temp_root, ignore_errors=True)

    summarize(doc_results, "SRS 文档", skip_stage_names={REFERENCE_STAGE_NAME})
    summarize(unit_results, "语义单元")

    # 如果启用了合并迭代通过项功能，执行合并
    if not args.dry_run and args.merge_iter_passes:
        print("\n开始合并迭代阶段通过项...")
        if args.mode in ("all", "docs"):
            merge_iter_evaluations(doc_eval_root, normalize_doc_stage_name, is_units=False)
        if args.mode in ("all", "units"):
            merge_iter_evaluations(unit_eval_root, normalize_unit_stage_name, is_units=True)
        print("迭代阶段合并完成\n")

    if not args.dry_run and not args.skip_summary_xlsx:
        summary_path = (
            Path(args.summary_xlsx).expanduser().resolve()
            if args.summary_xlsx
            else eval_output_dir / "数据汇总v2.xlsx"
        )
        generate_summary_workbook(
            summary_path=summary_path,
            doc_eval_root=doc_eval_root,
            unit_eval_root=unit_eval_root,
            outputs_dir=outputs_dir,
        )


def execute_jobs(jobs: Sequence[StageJob], max_parallel: int) -> List[tuple[Stage, int]]:
    """执行阶段任务，可配置并行度"""
    if not jobs:
        return []
    if max_parallel <= 1 or len(jobs) == 1:
        return [run_stage(*job) for job in jobs]

    results: List[tuple[Stage, int]] = []
    with ThreadPoolExecutor(max_workers=max_parallel) as executor:
        future_map = {executor.submit(run_stage, *job): job[0] for job in jobs}
        for future in as_completed(future_map):
            results.append(future.result())
    return results


if __name__ == "__main__":
    main()

