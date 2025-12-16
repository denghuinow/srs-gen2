#!/usr/bin/env python3
"""
使用指定权重配置重新计算加权得分并生成数据汇总Excel
"""
import json
import sys
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

# 导入eval_batch_outputs中的相关函数
sys.path.insert(0, str(Path(__file__).parent))
from eval_batch_outputs import (
    DocumentScore,
    StageEvalSummary,
    DIMENSION_ORDER,
    REFERENCE_STAGE_NAME,
    collect_document_names,
    format_score,
    format_ratio,
    normalize_doc_stage_name,
    normalize_unit_stage_name,
    safe_mean,
    stage_sort_key,
    write_dimension_sheet,
    write_stage_score_sheet,
    write_stage_summary_sheet,
    write_time_sheet,
    load_time_stats,
    build_stage_score_rows,
    build_stage_summary_rows,
    build_dimension_rows,
)

CATEGORIES = [
    "FUNCTIONAL",
    "BUSINESS_FLOW",
    "BOUNDARY",
    "EXCEPTION",
    "DATA_STATE",
    "CONSISTENCY_RULE",
]


def calculate_weighted_score(category_scores: Dict[str, Optional[float]], weights: Dict[str, float]) -> float:
    """使用指定权重计算加权得分"""
    weighted = sum(weights[cat] * (category_scores.get(cat, 0.0) or 0.0) for cat in weights)
    scaled = min(100.0, weighted * 2.0)
    return round(scaled, 2)


def load_stage_evaluations_with_weights(
    stage_root: Path, 
    name_normalizer,
    weights: Dict[str, float]
) -> List[StageEvalSummary]:
    """从评估输出目录加载阶段数据，并使用指定权重重新计算加权得分"""
    
    if not stage_root.exists():
        return []
    stages: List[StageEvalSummary] = []
    for directory in sorted(stage_root.iterdir(), key=lambda p: p.name.lower()):
        if not directory.is_dir():
            continue
        documents: Dict[str, DocumentScore] = {}
        for json_file in sorted(directory.glob("*_evaluation.json")):
            try:
                data = json.loads(json_file.read_text(encoding="utf-8"))
            except json.JSONDecodeError as exc:
                print(f"⚠ 无法解析 {json_file}: {exc}")
                continue
            doc_name = json_file.name.replace("_evaluation.json", "")
            scores = data.get("scores", {})
            category_scores = {
                cat: info.get("score")
                for cat, info in (scores.get("categories") or {}).items()
            }
            
            # 使用新权重重新计算加权得分
            new_weighted = calculate_weighted_score(category_scores, weights)
            
            documents[doc_name] = DocumentScore(
                weighted=new_weighted,  # 使用新计算的加权得分
                voting=scores.get("voting_score"),
                average=scores.get("average_score"),
                categories=category_scores,
            )
        if documents:
            stages.append(
                StageEvalSummary(
                    raw_name=directory.name,
                    display_name=name_normalizer(directory.name),
                    documents=documents,
                )
            )
    stages.sort(key=lambda stage: stage_sort_key(stage.display_name))
    return stages


def generate_summary_workbook_with_weights(
    summary_path: Path,
    doc_eval_root: Path,
    unit_eval_root: Path,
    outputs_dir: Path,
    weights: Dict[str, float],
    weight_name: str,
) -> None:
    """使用指定权重生成数据汇总Excel"""
    
    try:
        from openpyxl import Workbook
    except ImportError:
        print("⚠ 未安装 openpyxl，跳过汇总 Excel 生成")
        return
    
    print(f"使用权重配置 '{weight_name}' 生成数据汇总...")
    print(f"权重配置: {weights}")
    
    doc_stages = load_stage_evaluations_with_weights(doc_eval_root, normalize_doc_stage_name, weights)
    unit_stages = load_stage_evaluations_with_weights(unit_eval_root, normalize_unit_stage_name, weights)
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    
    # 添加权重信息工作表
    weight_sheet = wb.active
    weight_sheet.title = "权重配置"
    weight_sheet.append(["权重配置名称", weight_name])
    weight_sheet.append([])
    weight_sheet.append(["维度", "权重"])
    for cat in CATEGORIES:
        weight_sheet.append([cat, weights.get(cat, 0.0)])
    weight_sheet.append(["总和", sum(weights.values())])
    
    doc_sheet = wb.create_sheet("文档逐阶段")
    write_stage_score_sheet(doc_sheet, doc_stages, "阶段/文档")
    
    doc_summary_sheet = wb.create_sheet("文档阶段统计")
    write_stage_summary_sheet(doc_summary_sheet, doc_stages)
    
    doc_dimension_sheet = wb.create_sheet("文档维度")
    write_dimension_sheet(doc_dimension_sheet, doc_stages)
    
    unit_sheet = wb.create_sheet("语义单元逐阶段")
    write_stage_score_sheet(unit_sheet, unit_stages, "阶段", include_reference_row=False)
    
    unit_summary_sheet = wb.create_sheet("语义单元阶段统计")
    write_stage_summary_sheet(unit_summary_sheet, unit_stages, include_reference_row=False)
    
    unit_dimension_sheet = wb.create_sheet("语义单元维度")
    write_dimension_sheet(unit_dimension_sheet, unit_stages, include_reference_row=False)
    
    time_sheet = wb.create_sheet("生成耗时")
    time_rows = load_time_stats(outputs_dir / "耗时统计-总耗时.csv")
    write_time_sheet(time_sheet, time_rows, outputs_dir / "耗时统计-总耗时.csv")
    
    wb.save(summary_path)
    print(f"汇总 Excel 已写入：{summary_path}")


def main():
    # 方案12权重配置
    weights_12 = {
        "FUNCTIONAL": 0.80,
        "BUSINESS_FLOW": 0.05,
        "BOUNDARY": 0.02,
        "EXCEPTION": 0.06,
        "DATA_STATE": 0.05,
        "CONSISTENCY_RULE": 0.02,
    }
    
    # 方案13权重配置
    weights_13 = {
        "FUNCTIONAL": 0.55,
        "BUSINESS_FLOW": 0.05,
        "BOUNDARY": 0.02,
        "EXCEPTION": 0.20,
        "DATA_STATE": 0.15,
        "CONSISTENCY_RULE": 0.03,
    }
    
    # 路径配置
    eval_root = Path("/root/project/srs/srs-gen2/eval_reports/en_iter10")
    doc_eval_root = eval_root / "documents"
    unit_eval_root = eval_root / "units"
    outputs_dir = Path("/root/project/srs/srs-gen2/output/en_iter10")
    
    # 生成方案12的数据汇总
    summary_12_path = eval_root / "数据汇总v2_方案12.xlsx"
    generate_summary_workbook_with_weights(
        summary_12_path,
        doc_eval_root,
        unit_eval_root,
        outputs_dir,
        weights_12,
        "方案12: 几乎完全依赖FUNCTIONAL",
    )
    
    # 生成方案13的数据汇总
    summary_13_path = eval_root / "数据汇总v2_方案13.xlsx"
    generate_summary_workbook_with_weights(
        summary_13_path,
        doc_eval_root,
        unit_eval_root,
        outputs_dir,
        weights_13,
        "方案13: 高FUNCTIONAL+高EXCEPTION+DATA_STATE",
    )
    
    print("\n完成！已生成以下文件：")
    print(f"  - {summary_12_path}")
    print(f"  - {summary_13_path}")


if __name__ == "__main__":
    main()

